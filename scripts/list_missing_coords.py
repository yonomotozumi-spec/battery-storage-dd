#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
座標が未取得のまま残っている変電所の、座標調査の優先リストを作る。

OSM（Overpass）からの取得は fetch_substation_coords.py で掘り尽くしており、
残る未取得行はOSMに登録が無い。埋めるには各社の系統情報公表資料に当たるしかないため、
どこから手を付けるかを決めるための作業リストを出す。

優先度は「接続検討を申し込む価値」＝系統スコアのランクに合わせる。
ランクS/A（系統スコア48点以上）かつゲート除外でない行だけが対象。
S/A以外は座標が付いても申し込み対象になりにくいので載せない。

  優先S/A     : 販売対象エリアのS/A。系統スコアの降順。
  劣後S/A     : --defer で指定したエリア（販売が難しいエリア）のS/A。
  都道府県別   : 何県分の資料に当たれば何件埋まるかの集計。着手順の判断用。

緯度・経度は空欄で出力する。調べた座標を書き込んで
merge_coords.py で本体xlsxへ戻す運用を想定している。

使い方:
  python scripts/list_missing_coords.py \
    --in 全国変電所リスト_66kV以上_座標追加.xlsx \
    --out docs/座標未取得_優先リスト.xlsx --defer 四国 --defer 北陸
"""
import argparse
import collections
import urllib.parse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from score_substations import read_substations
import substation_scoring as S

HEADER_FILL = PatternFill("solid", fgColor="FCE4D6")
DEFER_FILL = PatternFill("solid", fgColor="EDEDED")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # 記入する列
TITLE_FONT = Font(bold=True, size=13)
HDR_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# read_substations が返す列の並び
FIELDS = ["no", "area", "name", "pref", "vmax", "vsec", "lat", "lon", "acc", "cap",
          "opcap", "flow", "avail", "avail_up", "n1", "curtail", "addr"]

# エリア → 座標を探しに行く先（一般送配電事業者。系統情報の公表元）
AREA_OPERATOR = {
    "北海道": "北海道電力ネットワーク",
    "東北": "東北電力ネットワーク",
    "東京": "東京電力パワーグリッド",
    "中部": "中部電力パワーグリッド",
    "北陸": "北陸電力送配電",
    "関西": "関西電力送配電",
    "中国": "中国電力ネットワーク",
    "四国": "四国電力送配電",
    "九州": "九州電力送配電",
    "沖縄": "沖縄電力",
    "J-POWER": "電源開発",
}

COLUMNS = [
    ("優先順位", 8), ("ランク", 7), ("系統スコア", 10), ("エリア", 8), ("変電所名", 20),
    ("都道府県", 11), ("最大電圧\nkV", 9), ("二次電圧\nkV", 9), ("空容量\n上位系MW", 11),
    ("空容量\n当該MW", 11), ("N-1電制", 9), ("出力制御", 9), ("問い合わせ先", 22),
    ("検索リンク", 12), ("緯度", 11), ("経度", 11), ("出典メモ", 22),
]


def maps_query(row):
    """変電所名だけでは同名が多いので、都道府県と事業者名を添えた検索リンクにする。"""
    terms = [t for t in (row["pref"], AREA_OPERATOR.get(row["area"], ""), row["name"]) if t]
    return "https://www.google.com/maps/search/" + urllib.parse.quote(" ".join(terms))


def load_targets(src, sheet):
    """S/Aかつゲート除外でない、座標未取得の行を返す。"""
    rows = [dict(zip(FIELDS, r)) for r in read_substations(src, sheet)]
    targets = []
    for r in rows:
        if isinstance(r["lat"], (int, float)):
            continue
        sc = S.score_row(r["area"], r["vsec"], r["opcap"], r["flow"],
                         r["avail"], r["avail_up"], r["n1"], r["curtail"])
        if sc["gate"] == "除外" or sc["rank"] not in ("S", "A"):
            continue
        r["rank"] = sc["rank"]
        r["grid"] = sc["grid_score"]
        targets.append(r)
    # 系統スコアの高い順。同点はエリア→名称で安定させる
    targets.sort(key=lambda r: (-r["grid"], r["area"], str(r["name"])))
    return rows, targets


def write_list_sheet(ws, title, targets, defer=False):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for c, (h, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(2, c, h)
        cell.fill = DEFER_FILL if defer else HEADER_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, r in enumerate(targets):
        row = 3 + i
        values = [i + 1, r["rank"], r["grid"], r["area"], r["name"], r["pref"],
                  r["vmax"], r["vsec"], r["avail_up"], r["avail"], r["n1"], r["curtail"],
                  AREA_OPERATOR.get(r["area"], ""), None, None, None, None]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.border = BORDER
        link = ws.cell(row, 14, "地図で探す")
        link.hyperlink = maps_query(r)
        link.font = Font(color="0563C1", underline="single")
        for c in (15, 16, 17):        # 緯度・経度・出典メモは記入用
            ws.cell(row, c).fill = INPUT_FILL
    ws.freeze_panes = "A3"
    return len(targets)


def write_pref_sheet(ws, targets, defer_areas):
    ws["A1"] = "都道府県別の残件数（資料に当たる順番の判断用。都道府県不明の行はエリア名で集計）"
    ws["A1"].font = TITLE_FONT
    headers = ["都道府県/エリア", "エリア", "問い合わせ先", "S/A残件", "うちS", "扱い"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = HEADER_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    widths = [18, 10, 24, 10, 8, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    agg = collections.defaultdict(lambda: {"n": 0, "s": 0})
    for r in targets:
        key = (r["pref"] or f'({r["area"]}・都道府県不明)', r["area"])
        agg[key]["n"] += 1
        agg[key]["s"] += 1 if r["rank"] == "S" else 0

    for i, ((pref, area), v) in enumerate(sorted(agg.items(), key=lambda kv: -kv[1]["n"])):
        row = 3 + i
        deferred = area in defer_areas
        values = [pref, area, AREA_OPERATOR.get(area, ""), v["n"], v["s"],
                  "劣後" if deferred else "優先"]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row, c, val)
            cell.border = BORDER
            if deferred:
                cell.fill = DEFER_FILL
    ws.freeze_panes = "A3"


def main():
    ap = argparse.ArgumentParser(description="座標未取得のS/A変電所の優先リストを出力する")
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--out", dest="out", required=True)
    ap.add_argument("--list-sheet", default="全国変電所リスト")
    ap.add_argument("--defer", action="append", default=[],
                    help="劣後させるエリア名。複数指定可（例: --defer 四国 --defer 北陸）")
    args = ap.parse_args()

    all_rows, targets = load_targets(args.src, args.list_sheet)
    defer_areas = set(args.defer)
    priority = [r for r in targets if r["area"] not in defer_areas]
    deferred = [r for r in targets if r["area"] in defer_areas]

    wb = Workbook()
    ws = wb.active
    ws.title = "優先S_A"
    write_list_sheet(ws, "座標未取得のS/A変電所（優先）　黄色=調べた座標の記入欄", priority)
    if deferred:
        write_list_sheet(wb.create_sheet("劣後S_A"),
                         f"座標未取得のS/A変電所（劣後: {'・'.join(sorted(defer_areas))}）",
                         deferred, defer=True)
    write_pref_sheet(wb.create_sheet("都道府県別"), targets, defer_areas)
    wb.save(args.out)

    missing = sum(1 for r in all_rows if not isinstance(r["lat"], (int, float)))
    print(f"座標未取得 {missing}件 / うちS/A(ゲート除外を除く) {len(targets)}件")
    print(f"  優先 {len(priority)}件 / 劣後 {len(deferred)}件")
    print(f"出力: {args.out}")


if __name__ == "__main__":
    main()
