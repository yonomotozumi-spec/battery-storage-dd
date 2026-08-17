#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
優先リスト（list_missing_coords.py の出力）に手で書き込んだ座標を、本体xlsxへ戻す。

OSMに無い変電所は各社の系統情報公表資料に当たるしかなく、その調査結果を
優先リストの緯度・経度欄に書き込んで運用する。このスクリプトはその記入分だけを
本体xlsxへ反映し、何を採用し何を弾いたかをレポートに残す。

突き合わせは No.（本体xlsxの通し番号）で行う。エリア＋変電所名は13組が重複しており
キーにできないため、優先リストのNo.列は消さないこと。

【採用しないもの】座標の取り違えは、そのまま距離スコアの誤りになる
  - 既に座標がある行（--force で上書きできるが既定では触らない）
  - 日本の範囲外の値、緯度経度の取り違えと思われる値
  - そのエリアの実測分布から外れた値（fetch_substation_coords.py と同じ判定）

使い方:
  python scripts/merge_coords.py \
    --in 全国変電所リスト_66kV以上_座標追加.xlsx \
    --filled docs/座標未取得_優先リスト.xlsx \
    --out 全国変電所リスト_66kV以上_座標追加.xlsx --report merge_report.csv
"""
import argparse
import csv

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from fetch_substation_coords import area_ranges, in_area, read_rows

ACC_MANUAL = "公表資料(手動)"
MANUAL_FILL = PatternFill("solid", fgColor="DDEBF7")   # 青=手動で調べた座標

# 日本の外接範囲。これを外れる値は入力ミスとみなす（南鳥島・与那国島を含む余裕を持たせた範囲）
JP_LAT = (20.0, 46.0)
JP_LON = (122.0, 154.0)


def read_filled(path, sheets):
    """優先リストから、緯度・経度が記入された行を読む。

    戻り値: [{"no", "name", "lat", "lon", "memo", "sheet", "row"}]
    """
    wb = load_workbook(path, data_only=True)
    out = []
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(2, c).value or "").replace("\n", "")
            if h:
                headers[h] = c
        for token in ("No.", "緯度", "経度"):
            if token not in headers:
                raise SystemExit(f"シート'{sheet}'に列 '{token}' がありません: {list(headers)}")
        for r in range(3, ws.max_row + 1):
            lat = ws.cell(r, headers["緯度"]).value
            lon = ws.cell(r, headers["経度"]).value
            if lat is None and lon is None:
                continue
            out.append({
                "no": ws.cell(r, headers["No."]).value,
                "name": ws.cell(r, headers["変電所名"]).value if "変電所名" in headers else "",
                "lat": lat, "lon": lon,
                "memo": ws.cell(r, headers["出典メモ"]).value if "出典メモ" in headers else "",
                "sheet": sheet, "row": r,
            })
    wb.close()
    return out


def validate(entry, target, box, force):
    """採用可否を (可否, 理由) で返す。"""
    if target is None:
        return False, "No.が本体xlsxに無い"
    # No.の書き間違いや行ズレで別の変電所に書き込むのを防ぐ
    if entry["name"] and str(entry["name"]).strip() != target["name"]:
        return False, f"No.と変電所名が一致しない（本体は「{target['name']}」）"
    lat, lon = entry["lat"], entry["lon"]
    if not isinstance(lat, (int, float)) or not isinstance(lon, (int, float)):
        return False, "緯度・経度が数値でない（片方だけの記入も不可）"
    if not (JP_LAT[0] <= lat <= JP_LAT[1] and JP_LON[0] <= lon <= JP_LON[1]):
        # 緯度と経度を逆に書いた場合もここで落ちる
        return False, f"日本の範囲外（緯度経度の取り違え？）: {lat},{lon}"
    if isinstance(target["lat"], (int, float)) and not force:
        return False, "既に座標がある（上書きするなら --force）"
    if not in_area(box, target["area"], lat, lon):
        return False, f"{target['area']}エリアの実測分布から外れている"
    return True, "採用"


def main():
    ap = argparse.ArgumentParser(description="優先リストに記入した座標を本体xlsxへ反映する")
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--filled", required=True, help="座標を記入した優先リストのxlsx")
    ap.add_argument("--out", dest="out", required=True)
    ap.add_argument("--report", help="採用・不採用の一覧CSV")
    ap.add_argument("--list-sheet", default="全国変電所リスト")
    ap.add_argument("--filled-sheets", default="優先S_A,劣後S_A",
                    help="読み込む優先リストのシート名（カンマ区切り）")
    ap.add_argument("--force", action="store_true", help="既に座標がある行も上書きする")
    args = ap.parse_args()

    rows, cols, _ = read_rows(args.src, args.list_sheet)
    by_no = {r["no"]: r for r in rows if r["no"] is not None}
    box = area_ranges(rows)
    entries = read_filled(args.filled, [s.strip() for s in args.filled_sheets.split(",")])
    if not entries:
        raise SystemExit("優先リストに記入された座標がありません")

    wb = load_workbook(args.src)
    ws = wb[args.list_sheet]

    applied, skipped, report = 0, 0, []
    for e in entries:
        target = by_no.get(e["no"])
        ok, why = validate(e, target, box, args.force)
        if ok:
            ws.cell(target["row"], cols["lat"]).value = e["lat"]
            ws.cell(target["row"], cols["lon"]).value = e["lon"]
            ws.cell(target["row"], cols["acc"]).value = ACC_MANUAL
            for c in (cols["lat"], cols["lon"], cols["acc"]):
                ws.cell(target["row"], c).fill = MANUAL_FILL
            applied += 1
        else:
            skipped += 1
        report.append([e["sheet"], e["row"], e["no"], e["name"] or (target or {}).get("name", ""),
                       e["lat"], e["lon"], "採用" if ok else "不採用", why, e["memo"] or ""])

    wb.save(args.out)
    print(f"記入 {len(entries)}件 → 反映 {applied}件 / 不採用 {skipped}件")
    print(f"出力: {args.out}")
    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["シート", "行", "No.", "変電所名", "緯度", "経度", "判定", "理由", "出典メモ"])
            w.writerows(report)
        print(f"レポート: {args.report}")
    if applied:
        print("※ 反映後は score_substations.py と build_substation_web_data.py の再実行が必要です。")


if __name__ == "__main__":
    main()
