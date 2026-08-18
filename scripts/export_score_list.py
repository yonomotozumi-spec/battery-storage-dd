#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
変電所スコア一覧 xlsx 出力スクリプト（系統用蓄電池・高圧2MW前提 v1）

score_substations.py が出力する「スコアリング」タブは元リストの並び（エリア順）で、
スコアは数式のためExcelで開かないと値が見えない。本スクリプトはその値版として、
全変電所を **系統スコアの高い順に並べた一覧** を出力する。
配点は substation_scoring.py（xlsx版・Web版と共通の単一ソース）をそのまま使う。

  スコア一覧 : 全行をランク→系統スコア降順で並べた値ベースの一覧。
               オートフィルタ・ウィンドウ枠固定つき。ランクS/Aは黄色、除外はグレー。
               座標がある行はGoogleマップリンク列から地図を開ける。
  集計       : ランク別件数・エリア×ランクのクロス集計・座標カバー率。
  スコア基準 : 配点定義（S1〜S6）・エリア点・ランク閾値・前提メモ。

使い方:
  python scripts/export_score_list.py \
    --in "data/全国変電所リスト_66kV以上_座標追加.xlsx" \
    --out "docs/変電所スコア一覧_全国66kV以上.xlsx"
"""
import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import substation_scoring as S
from score_substations import read_substations

HEADER_FILL = PatternFill("solid", fgColor="FCE4D6")
RANK_FILL = PatternFill("solid", fgColor="FFF2A8")    # ランクS/A
GATE_FILL = PatternFill("solid", fgColor="E7E6E6")    # ゲート除外
HDR_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SMALL = Font(size=10)
LINK_FONT = Font(size=10, color="0563C1", underline="single")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical="center")

# 並び順: ランク → 系統スコア降順 → エリア → 変電所名
RANK_ORDER = {"S": 0, "A": 1, "B": 2, "C": 3, "データ無": 4, "除外": 5}

COLUMNS = [
    ("順位", 6), ("ランク", 7), ("系統スコア(80)", 12), ("エリア", 8), ("変電所名", 26),
    ("都道府県", 10), ("所在地(参考)", 20), ("最大電圧kV", 10), ("二次電圧kV", 10),
    ("空容量_採用MW", 12), ("空容量_当該MW", 12), ("空容量_上位系MW", 13),
    ("N-1電制", 14), ("出力制御", 9), ("運用容量MW", 10), ("予想潮流MW", 10), ("潮流率", 8),
    ("S1空容量(35)", 10), ("S2 N-1(10)", 10), ("S3潮流(15)", 10), ("S4出制(5)", 9),
    ("S5配変(10)", 10), ("S6エリア(5)", 10),
    ("緯度", 10), ("経度", 10), ("座標精度", 12), ("地図", 8),
    ("実績メモ", 30), ("元リストNo.", 10),
]


def num(v):
    """数値ならfloat/intで返す。文字列や空欄は None。"""
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        return None
    return v


def build_rows(src_rows):
    """入力行にスコアを付け、ランク→スコア降順に並べ替えた一覧を返す。"""
    out = []
    for (no, area, name, pref, vmax, vsec, lat, lon, acc, _cap, opcap, flow,
         avail, avail_up, n1, curtail, addr) in src_rows:
        sc = S.score_row(area, vsec, opcap, flow, avail, avail_up, n1, curtail)
        eff = S.effective_capacity(avail, avail_up)
        cap, fl = num(opcap), num(flow)
        ratio = abs(fl) / cap if (cap and fl is not None) else None
        out.append({
            "no": no, "area": area, "name": name, "pref": pref, "addr": addr,
            "vmax": vmax, "vsec": vsec, "eff": eff, "avail": avail, "avail_up": avail_up,
            "n1": n1, "curtail": curtail, "opcap": opcap, "flow": flow, "ratio": ratio,
            "lat": num(lat), "lon": num(lon), "acc": acc,
            "memo": S.RESULTS_MEMO.get(S.results_key(area, name), ""),
            **sc,
        })
    out.sort(key=lambda r: (RANK_ORDER.get(r["rank"], 9), -r["grid_score"],
                            str(r["area"]), str(r["name"] or "")))
    return out


def build_list_sheet(ws, rows):
    ws["A1"] = ("全国変電所スコア一覧（高圧2MW BESS・系統スコア80点の高い順）"
                "　黄色=ランクS/A ／ グレー=ゲート除外")
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"全{len(rows)}件 ｜ 配点: 変電所スコアリング v1（S1空容量35/S2 N-1電制10/S3潮流15/"
                "S4出力制御5/S5配変10/S6エリア5）｜ 土地確定後にS7近接20点を加えて100点満点で最終評価")
    ws["A2"].font = SMALL

    hdr_row = 3
    for c, (h, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(hdr_row, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, r in enumerate(rows, 1):
        row = hdr_row + i
        values = [
            i, r["rank"], r["grid_score"], r["area"], r["name"], r["pref"], r["addr"],
            r["vmax"], r["vsec"], r["eff"], r["avail"], r["avail_up"],
            r["n1"], r["curtail"], r["opcap"], r["flow"], r["ratio"],
            r["s1"], r["s2"], r["s3"], r["s4"], r["s5"], r["s6"],
            r["lat"], r["lon"], r["acc"], None, r["memo"], r["no"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.border = BORDER
            cell.alignment = CENTER if c <= 3 or 8 <= c <= 26 else WRAP
        ws.cell(row, 17).number_format = "0.0%"
        for c in (24, 25):
            ws.cell(row, c).number_format = "0.00000"
        if r["lat"] is not None and r["lon"] is not None:
            link = ws.cell(row, 27, "地図")
            link.hyperlink = f"https://www.google.com/maps?q={r['lat']},{r['lon']}"
            link.font = LINK_FONT
            link.alignment = CENTER
        if r["rank"] in ("S", "A"):
            fill = RANK_FILL
        elif r["rank"] == "除外":
            fill = GATE_FILL
        else:
            fill = None
        if fill:
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(row, c).fill = fill

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(COLUMNS))}{hdr_row + len(rows)}"
    ws.freeze_panes = f"F{hdr_row + 1}"


def build_summary_sheet(ws, rows):
    ws["A1"] = "集計（ランク別・エリア別）"
    ws["A1"].font = TITLE_FONT

    ranks = ["S", "A", "B", "C", "データ無", "除外"]
    areas = [a for a, _, _ in S.AREA_POINTS if any(r["area"] == a for r in rows)]
    areas += sorted({str(r["area"]) for r in rows} - set(areas))

    r0 = 3
    ws.cell(r0, 1, "ランク別件数").font = HDR_FONT
    for c, h in enumerate(["ランク", "件数", "構成比", "座標あり", "説明"], 1):
        cell = ws.cell(r0 + 1, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    desc = {
        "S": f"系統スコア{S.P['rank_s']['thr']}点以上・最優先で接続検討",
        "A": f"{S.P['rank_a']['thr']}〜{S.P['rank_s']['thr'] - 1}点・優先候補",
        "B": f"{S.P['rank_b']['thr']}〜{S.P['rank_a']['thr'] - 1}点・条件次第",
        "C": f"{S.P['rank_b']['thr'] - 1}点以下",
        "データ無": "空容量が未公表（OSM等の座標のみ）",
        "除外": "空容量≦0かつN-1電制不可（上位系増強リスク）",
    }
    for i, rk in enumerate(ranks):
        sub = [r for r in rows if r["rank"] == rk]
        with_xy = sum(1 for r in sub if r["lat"] is not None and r["lon"] is not None)
        vals = [rk, len(sub), (len(sub) / len(rows) if rows else 0), with_xy, desc[rk]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r0 + 2 + i, c, v); cell.border = BORDER
            cell.alignment = WRAP if c == 5 else CENTER
        ws.cell(r0 + 2 + i, 3).number_format = "0.0%"
        if rk in ("S", "A"):
            for c in range(1, 6):
                ws.cell(r0 + 2 + i, c).fill = RANK_FILL
    total_row = r0 + 2 + len(ranks)
    with_xy = sum(1 for r in rows if r["lat"] is not None and r["lon"] is not None)
    for c, v in enumerate(["合計", len(rows), 1.0, with_xy,
                           f"座標カバー率 {with_xy / len(rows):.1%}（距離検索が可能な件数）"], 1):
        cell = ws.cell(total_row, c, v)
        cell.border = BORDER; cell.font = HDR_FONT
        cell.alignment = WRAP if c == 5 else CENTER
    ws.cell(total_row, 3).number_format = "0.0%"

    r1 = total_row + 2
    ws.cell(r1, 1, "エリア×ランク クロス集計").font = HDR_FONT
    head = ["エリア"] + ranks + ["合計", "S/A計", "S6エリア点"]
    for c, h in enumerate(head, 1):
        cell = ws.cell(r1 + 1, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for i, area in enumerate(areas):
        sub = [r for r in rows if str(r["area"]) == area]
        counts = [sum(1 for r in sub if r["rank"] == rk) for rk in ranks]
        sa = counts[0] + counts[1]
        vals = [area] + counts + [len(sub), sa, S.AREA_POINT_MAP.get(area, 0)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r1 + 2 + i, c, v)
            cell.border = BORDER; cell.alignment = CENTER

    for c, w in enumerate([12, 10, 10, 10, 10, 10, 10, 10, 12, 46], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_basis_sheet(ws):
    ws["A1"] = "スコア基準（系統用蓄電池・高圧2MW前提）v1"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "定義の単一ソース: scripts/substation_scoring.py（xlsx版・Web版と共通）"
    ws["A2"].font = SMALL

    r = 4
    ws.cell(r, 1, "変数・配点定義").font = HDR_FONT
    r += 1
    for c, h in enumerate(["記号", "変数", "配点", "閾値・ルール", "根拠(実績)"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for row in S.SCORE_DEFS:
        r += 1
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v); cell.border = BORDER; cell.alignment = WRAP

    r += 2
    ws.cell(r, 1, "S6 エリア点").font = HDR_FONT
    r += 1
    for c, h in enumerate(["エリア", "点数", "根拠"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for area, pts, why in S.AREA_POINTS:
        r += 1
        for c, v in enumerate([area, pts, why], 1):
            cell = ws.cell(r, c, v); cell.border = BORDER; cell.alignment = WRAP

    r += 2
    ws.cell(r, 1, "ランク閾値・ゲート").font = HDR_FONT
    for label, text in [
        ("ランクS", f"系統スコア {S.P['rank_s']['thr']} 点以上"),
        ("ランクA", f"系統スコア {S.P['rank_a']['thr']} 〜 {S.P['rank_s']['thr'] - 1} 点"),
        ("ランクB", f"系統スコア {S.P['rank_b']['thr']} 〜 {S.P['rank_a']['thr'] - 1} 点"),
        ("ランクC", f"系統スコア {S.P['rank_b']['thr'] - 1} 点以下"),
        ("データ無", "空容量が当該設備・上位系考慮ともに未公表"),
        ("ゲート除外", "空容量(上位系考慮)≦0 かつ N-1電制≠可（上位系増強リスク。実績: 喜多方=工期10年2ヶ月）"),
    ]:
        r += 1
        ws.cell(r, 1, label).font = HDR_FONT
        ws.cell(r, 1).border = BORDER
        cell = ws.cell(r, 2, text); cell.alignment = WRAP; cell.border = BORDER

    r += 2
    ws.cell(r, 1, "前提・使い方").font = HDR_FONT
    for note in S.NOTES:
        r += 1
        cell = ws.cell(r, 1, note); cell.alignment = WRAP; cell.font = SMALL

    for c, w in enumerate([14, 24, 8, 44, 40], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def main():
    ap = argparse.ArgumentParser(description="変電所スコア一覧xlsx（スコア降順・値ベース）を出力")
    ap.add_argument("--in", dest="src", required=True, help="全国変電所リスト（66kV以上）のxlsxパス")
    ap.add_argument("--out", dest="out", required=True, help="出力xlsxパス")
    ap.add_argument("--list-sheet", default="全国変電所リスト", help="入力の変電所リストのシート名")
    args = ap.parse_args()

    src_rows = read_substations(args.src, args.list_sheet)
    print(f"読込: {len(src_rows)}行")
    rows = build_rows(src_rows)

    wb = Workbook()
    build_list_sheet(wb.active, rows)
    wb.active.title = "スコア一覧"
    build_summary_sheet(wb.create_sheet("集計"), rows)
    build_basis_sheet(wb.create_sheet("スコア基準"))
    wb.save(args.out)

    tally = {rk: sum(1 for r in rows if r["rank"] == rk) for rk in RANK_ORDER}
    print("出力: " + args.out)
    print("ランク別: " + " / ".join(f"{k}={v}" for k, v in tally.items()))


if __name__ == "__main__":
    main()
