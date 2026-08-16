#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
変電所リストの座標未取得行に、OpenStreetMap（Overpass API）から実測座標を付与する。

全国変電所リスト（66kV以上）は、送配電事業者の公表データ由来の行に座標が無い。
このスクリプトは都道府県単位でOSMの変電所を取得し、名称を正規化して突き合わせ、
一致した行にだけ緯度・経度・座標精度を書き込む。

【方針】推測で座標を埋めない
  - 一致しなかった行は空欄のまま。都道府県の代表点などで埋めることはしない
    （Web版の近接スコアS7は500m〜5kmの刻みで判定するため、数十km誤差の推定値を
      入れると距離スコアが破綻する）。
  - 同名の変電所が全国に複数あるため、原則として都道府県内に限定して照合する。
    都道府県が不明な行は全国検索し、候補がちょうど1件のときだけ採用する。

【ネットワーク】
  Overpass APIへの直接アクセスが必要。遮断された環境では取得できないため、
  通信できる環境で `fetch` を実行して --cache-dir を持ち帰り、`apply --from-cache`
  でオフラインに反映する運用ができる。

使い方:
  # ① 取得（要ネットワーク）。都道府県ごとにJSONを cache/ に保存する
  python scripts/fetch_substation_coords.py fetch \
    --in 全国変電所リスト_66kV以上.xlsx --cache-dir osm_cache

  # ② 反映（ネットワーク不要）。キャッシュを突き合わせて新しいxlsxを書き出す
  python scripts/fetch_substation_coords.py apply \
    --in 全国変電所リスト_66kV以上.xlsx --cache-dir osm_cache \
    --out 全国変電所リスト_66kV以上_座標追加.xlsx --report coord_report.csv
"""
import argparse
import csv
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

OVERPASS_URL = "https://overpass-api.de/api/interpreter"
USER_AGENT = "battery-storage-dd substation coord fetcher (+https://github.com/yonomotozumi-spec/battery-storage-dd)"

ACC_OSM_NAME = "OSM実測(名称一致)"

# 全国名称検索の結果をエリアの実測分布と突き合わせるときの許容幅（度）と、
# 分布を作るのに最低限必要な参照点の数
AREA_BOX_MARGIN = 0.3
AREA_BOX_MIN_REF = 20
OSM_FILL = PatternFill("solid", fgColor="E2EFDA")  # 緑=OSM実測

# 変電所名の正規化で落とす事業者名プレフィックス
OPERATOR_PREFIXES = [
    "北海道電力ネットワーク", "東北電力ネットワーク", "東京電力パワーグリッド",
    "中部電力パワーグリッド", "北陸電力送配電", "関西電力送配電",
    "中国電力ネットワーク", "四国電力送配電", "九州電力送配電", "沖縄電力",
    "北海道電力", "東北電力", "東京電力", "中部電力", "北陸電力", "関西電力",
    "中国電力", "四国電力", "九州電力", "電源開発", "ＪＲ", "JR",
]
SUFFIXES = ["変電所", "開閉所", "変換所", "電力所", "S/S", "SS"]


def normalize_name(name):
    """変電所名の照合キー（strict）。事業者名・法人格・連番・空白は落とすが、
    施設種別（変電所／開閉所／変換所）は残す。

    種別まで落とすと「芳賀変電所」と「芳賀開閉所」のような別施設が同じキーになり、
    取り違えの原因になるため。種別を落とした緩いキーは loose_name() を使う。
    """
    if not name:
        return ""
    s = unicodedata.normalize("NFKC", str(name))  # ①→1 などもここで正規化される
    s = re.sub(r"\(株\)|\(有\)|株式会社|有限会社|㈱|㈲", "", s)
    for p in OPERATOR_PREFIXES:
        if s.startswith(p):
            s = s[len(p):]
            break
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[（(][^）)]*[）)]$", "", s)   # 末尾の括弧書き
    s = re.sub(r"[0-9]+$", "", s)              # 種別の後ろの連番（大野変電所①→…所1）
    kind = ""
    for suf in SUFFIXES:
        if s.endswith(suf):
            kind = suf
            s = s[: -len(suf)]
            break
    s = re.sub(r"[0-9]+$", "", s).strip()      # 種別の前の連番
    # 固有名が残らないものは照合に使わない。「6203変電所」のように名前自体が数字だと
    # 連番除去で固有名が消え、OSM側の名無しの「変電所」と同じキーになってしまうため。
    if not s:
        return ""
    return s + ("@" + kind if kind else "")


def loose_name(name):
    """施設種別も落とした緩い照合キー。候補が一意のときだけ採用する用途。"""
    return normalize_name(name).split("@")[0]


def overpass_query_pref(pref):
    """都道府県名(admin_level=4)のエリア内にある変電所をすべて取得するクエリ。"""
    return (
        "[out:json][timeout:180];"
        f'area["name"="{pref}"]["admin_level"="4"]->.a;'
        '(way["power"="substation"](area.a);'
        'node["power"="substation"](area.a);'
        'relation["power"="substation"](area.a););'
        "out center tags;"
    )


def overpass_query_japan_names(names):
    """名称の完全一致（正規表現）で全国から取得するクエリ。都道府県不明行の補完用。"""
    pattern = "|".join(re.escape(n) for n in names)
    return (
        "[out:json][timeout:300];"
        f'area["name"="日本"]["admin_level"="2"]->.a;'
        f'(way["power"="substation"]["name"~"^({pattern})$"](area.a);'
        f'node["power"="substation"]["name"~"^({pattern})$"](area.a);'
        f'relation["power"="substation"]["name"~"^({pattern})$"](area.a););'
        "out center tags;"
    )


MAX_RETRY_WAIT = 60  # バックオフの上限(秒)。長時間の再試行でも待ち過ぎない


def overpass_fetch(query, retries=3, sleep=5):
    """Overpass APIを叩いてJSONを返す。429/504や接続断はバックオフして再試行。

    Overpassは共有IPからのアクセスを一時的に遮断することがあり、その間は
    接続が reset される。--retries を増やすと1プロセスのまま粘れる
    （取得済み都道府県はキャッシュに残るので、途中で落ちてもやり直しは効く）。
    """
    data = urllib.parse.urlencode({"data": query}).encode()
    req = urllib.request.Request(OVERPASS_URL, data=data, headers={"User-Agent": USER_AGENT})
    for attempt in range(1, retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=300) as r:
                return json.loads(r.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code in (429, 502, 503, 504) and attempt < retries:
                wait = min(sleep * (2 ** (attempt - 1)), MAX_RETRY_WAIT)
                print(f"  HTTP {e.code} → {wait}秒待って再試行 ({attempt}/{retries})", file=sys.stderr)
                time.sleep(wait)
                continue
            raise
        except (urllib.error.URLError, TimeoutError, ConnectionError) as e:
            if attempt < retries:
                wait = min(sleep * (2 ** (attempt - 1)), MAX_RETRY_WAIT)
                print(f"  接続失敗({e}) → {wait}秒待って再試行 ({attempt}/{retries})", file=sys.stderr)
                time.sleep(wait)
                continue
            raise
    raise RuntimeError("Overpass fetch failed")


def parse_elements(payload):
    """Overpassの応答から (strict索引, loose索引) を作る。値は候補のリスト。"""
    strict, loose = {}, {}
    for el in payload.get("elements", []):
        tags = el.get("tags") or {}
        name = tags.get("name:ja") or tags.get("name")
        if not name:
            continue
        if el.get("type") == "node":
            lat, lon = el.get("lat"), el.get("lon")
        else:
            c = el.get("center") or {}
            lat, lon = c.get("lat"), c.get("lon")
        if lat is None or lon is None:
            continue
        volt = 0
        for v in re.findall(r"\d+", str(tags.get("voltage", "")) or "0"):
            volt = max(volt, int(v))
        key = normalize_name(name)
        if not key:
            continue
        cand = {"name": name, "lat": round(float(lat), 6), "lon": round(float(lon), 6),
                "voltage": volt / 1000.0 if volt else 0}   # OSMはV単位 → kV
        strict.setdefault(key, []).append(cand)
        lkey = loose_name(name)
        if lkey:
            loose.setdefault(lkey, []).append(cand)
    return strict, loose


def resolve(cands, vmax_kv):
    """候補から1件に決める。決められなければ None を返して見送る。

    候補が複数のときは、リストの最大電圧(kV)と一致するOSM候補が
    ちょうど1件あるときだけ採用する（推測で選ばない）。
    """
    if not cands:
        return None, "一致なし"
    if len(cands) == 1:
        return cands[0], "一致"
    if vmax_kv:
        exact = [c for c in cands if c["voltage"] and abs(c["voltage"] - vmax_kv) < 0.5]
        if len(exact) == 1:
            return exact[0], "一致(電圧で特定)"
    return None, f"候補{len(cands)}件で特定不可"


def area_ranges(rows):
    """既にOSM実測座標がある行から、エリアごとの緯度経度レンジを作る。

    全国名称検索は同名の別変電所を掴みやすい（例: 関西エリアの「稲美変電所」に
    北海道のOSM座標が付く）。都道府県が不明でもエリアは必ず判っているため、
    そのエリアの実測分布から外れた候補は採用しない。
    """
    pts = {}
    for r in rows:
        if (isinstance(r["lat"], (int, float)) and isinstance(r["lon"], (int, float))
                and str(r["acc"] or "").startswith("OSM実測")):
            pts.setdefault(r["area"], []).append((r["lat"], r["lon"]))

    def pct(seq, f):
        return seq[min(len(seq) - 1, max(0, int(len(seq) * f)))]

    box = {}
    for area, ps in pts.items():
        if len(ps) < AREA_BOX_MIN_REF:
            continue   # 参照点が少ないエリア(J-POWER等)は分布が作れないので判定しない
        lats = sorted(p[0] for p in ps)
        lons = sorted(p[1] for p in ps)
        box[area] = (pct(lats, 0.005) - AREA_BOX_MARGIN, pct(lats, 0.995) + AREA_BOX_MARGIN,
                     pct(lons, 0.005) - AREA_BOX_MARGIN, pct(lons, 0.995) + AREA_BOX_MARGIN)
    return box


def in_area(box, area, lat, lon):
    """レンジを作れなかったエリアは判定材料が無いので True（採用を妨げない）。"""
    b = box.get(area)
    if b is None:
        return True
    return b[0] <= lat <= b[1] and b[2] <= lon <= b[3]


def read_rows(path, sheet_name):
    """(行番号, 変電所名, 都道府県, 既存緯度) のリストとヘッダー位置を返す。"""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    hdr = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "No.":
            hdr = row[0].row
            break
    if hdr is None:
        raise SystemExit("入力シートに 'No.' ヘッダー行が見つかりません")
    headers = [str(c.value or "").replace("\n", "") for c in ws[hdr]]

    def col(token):
        for i, h in enumerate(headers):
            if token in h:
                return i + 1
        raise SystemExit(f"ヘッダー '{token}' が見つかりません: {headers}")

    cols = {"name": col("変電所名"), "pref": col("都道府県"), "lat": col("緯度"),
            "lon": col("経度"), "acc": col("座標精度"), "area": col("エリア"),
            "vmax": col("最大電圧")}
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, cols["name"]).value
        if not name:
            continue
        pref = ws.cell(r, cols["pref"]).value
        vmax = ws.cell(r, cols["vmax"]).value
        rows.append({
            "row": r,
            "name": str(name),
            "pref": pref.strip() if isinstance(pref, str) else "",
            "area": ws.cell(r, cols["area"]).value or "",
            "vmax": float(vmax) if isinstance(vmax, (int, float)) else None,
            "lat": ws.cell(r, cols["lat"]).value,
            "lon": ws.cell(r, cols["lon"]).value,
            "acc": ws.cell(r, cols["acc"]).value,
        })
    wb.close()
    return rows, cols, hdr


def cache_path(cache_dir, key):
    safe = re.sub(r"[^\w\-]", "_", key)
    return os.path.join(cache_dir, f"{safe}.json")


def cmd_fetch(args):
    rows, _, _ = read_rows(args.src, args.list_sheet)
    missing = [r for r in rows if not isinstance(r["lat"], (int, float))]
    prefs = sorted({r["pref"] for r in missing if r["pref"]})
    print(f"座標未取得 {len(missing)}件 / 対象都道府県 {len(prefs)}件")
    os.makedirs(args.cache_dir, exist_ok=True)

    for i, pref in enumerate(prefs, 1):
        path = cache_path(args.cache_dir, pref)
        if os.path.exists(path) and not args.force:
            print(f"[{i}/{len(prefs)}] {pref}: キャッシュ済みのためスキップ")
            continue
        print(f"[{i}/{len(prefs)}] {pref}: 取得中…")
        payload = overpass_fetch(overpass_query_pref(pref), args.retries, args.retry_sleep)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        print(f"    {len(payload.get('elements', []))} 件取得")
        time.sleep(args.sleep)

    # 都道府県が空の行は名称で全国検索（採用は一意ヒット時のみ）
    nameless = sorted({r["name"] for r in missing if not r["pref"]})
    if nameless:
        path = cache_path(args.cache_dir, "_japan_by_name")
        if os.path.exists(path) and not args.force:
            print("全国名称検索: キャッシュ済みのためスキップ")
        else:
            print(f"全国名称検索: {len(nameless)}件の名称を{args.chunk}件ずつ照会")
            merged = {"elements": []}
            for i in range(0, len(nameless), args.chunk):
                chunk = nameless[i:i + args.chunk]
                payload = overpass_fetch(overpass_query_japan_names(chunk), args.retries, args.retry_sleep)
                merged["elements"].extend(payload.get("elements", []))
                print(f"    {i + len(chunk)}/{len(nameless)} 照会済み（累計{len(merged['elements'])}件）")
                time.sleep(args.sleep)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(merged, f, ensure_ascii=False)
    print(f"完了: {args.cache_dir}")


def load_cache(cache_dir):
    """{都道府県: (strict, loose)} と 全国分(strict, loose) を返す。"""
    by_pref, japan = {}, ({}, {})
    if not os.path.isdir(cache_dir):
        raise SystemExit(f"キャッシュディレクトリがありません: {cache_dir}")
    for fn in sorted(os.listdir(cache_dir)):
        if not fn.endswith(".json"):
            continue
        with open(os.path.join(cache_dir, fn), encoding="utf-8") as f:
            payload = json.load(f)
        parsed = parse_elements(payload)
        if fn == "_japan_by_name.json":
            japan = parsed
        else:
            by_pref[os.path.splitext(fn)[0]] = parsed
    return by_pref, japan


def cmd_apply(args):
    rows, cols, _ = read_rows(args.src, args.list_sheet)
    by_pref, japan = load_cache(args.cache_dir)

    wb = load_workbook(args.src)
    ws = wb[args.list_sheet]

    box = area_ranges(rows)
    filled, ambiguous, unmatched, off_area = 0, 0, 0, 0
    report = []
    for r in rows:
        if isinstance(r["lat"], (int, float)):
            continue
        if r["pref"] and r["pref"] in by_pref:
            strict, loose = by_pref[r["pref"]]
            scope = r["pref"]
        elif not r["pref"]:
            strict, loose = japan
            scope = "全国"
        else:
            strict, loose = {}, {}
            scope = r["pref"]

        # ① 施設種別まで一致するキーで照合 → ② 見つからなければ種別を外した緩いキー
        best, how = resolve(strict.get(normalize_name(r["name"]), []), r["vmax"])
        if best is None:
            l_cands = loose.get(loose_name(r["name"]), [])
            if len(l_cands) == 1:
                best, how = l_cands[0], "一致(種別ゆれ)"
            elif l_cands:
                how = f"候補{len(l_cands)}件で特定不可"

        # 全国検索は同名の別変電所を掴みうるので、エリアの実測分布から外れたら捨てる
        if best is not None and scope == "全国" and not in_area(box, r["area"], best["lat"], best["lon"]):
            off_area += 1
            report.append([r["row"], r["name"], r["pref"],
                           f"{how}(全国)だが{r['area']}エリア外のため不採用",
                           "", "", best["name"]])
            continue

        if best is None:
            if "特定不可" in how:
                ambiguous += 1
            else:
                unmatched += 1
            report.append([r["row"], r["name"], r["pref"], how, "", "", ""])
            continue

        ws.cell(r["row"], cols["lat"]).value = best["lat"]
        ws.cell(r["row"], cols["lon"]).value = best["lon"]
        ws.cell(r["row"], cols["acc"]).value = ACC_OSM_NAME
        for c in (cols["lat"], cols["lon"], cols["acc"]):
            ws.cell(r["row"], c).fill = OSM_FILL
        filled += 1
        report.append([r["row"], r["name"], r["pref"], f"{how}({scope})",
                       best["lat"], best["lon"], best["name"]])

    wb.save(args.out)
    print(f"付与: {filled}件 / 候補複数で特定不可: {ambiguous}件 / "
          f"エリア外で不採用: {off_area}件 / 一致なし: {unmatched}件")
    print(f"出力: {args.out}")
    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["行","変電所名","都道府県","結果","緯度","経度","OSM側の名称"])
            w.writerows(report)
        print(f"照合レポート: {args.report}")
    print("※ 一致しなかった行は空欄のままです（推定値では埋めません）。")


def main():
    ap = argparse.ArgumentParser(description="OSMから変電所座標を取得して未取得行に付与する")
    sub = ap.add_subparsers(dest="cmd", required=True)

    f = sub.add_parser("fetch", help="Overpass APIから都道府県ごとに取得してキャッシュする（要ネットワーク）")
    f.add_argument("--in", dest="src", required=True)
    f.add_argument("--cache-dir", required=True)
    f.add_argument("--list-sheet", default="全国変電所リスト")
    f.add_argument("--sleep", type=float, default=3.0, help="リクエスト間隔(秒)")
    f.add_argument("--chunk", type=int, default=50, help="名称検索の1リクエストあたり件数")
    f.add_argument("--force", action="store_true", help="キャッシュがあっても取り直す")
    f.add_argument("--retries", type=int, default=3,
                   help="1リクエストあたりの再試行回数。Overpassに遮断されがちな環境では増やす")
    f.add_argument("--retry-sleep", type=float, default=5.0,
                   help="再試行の初期待ち時間(秒)。以後2倍ずつ、最大%d秒" % MAX_RETRY_WAIT)
    f.set_defaults(func=cmd_fetch)

    a = sub.add_parser("apply", help="キャッシュを突き合わせてxlsxに座標を書き込む（ネットワーク不要）")
    a.add_argument("--in", dest="src", required=True)
    a.add_argument("--cache-dir", required=True)
    a.add_argument("--out", required=True)
    a.add_argument("--list-sheet", default="全国変電所リスト")
    a.add_argument("--report", help="照合結果のCSV出力先")
    a.set_defaults(func=cmd_apply)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
