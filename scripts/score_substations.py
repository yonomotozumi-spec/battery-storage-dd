#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
変電所スコアリング xlsx 生成スクリプト（系統用蓄電池・高圧2MW前提 v1）

全国変電所リスト（66kV以上・系統接続検討用）xlsx を読み込み、
接続検討の実績12件から校正した配点（変電所スコアリング v1 2026-07-28）で
全変電所を「接続検討を申し込む価値」の順に点数化した3タブのxlsxを出力する。

  スコアリング     : 全行のゲート判定・S1〜S6・系統スコア(80点)・ランク・実績メモ。
                     スコア列は全て数式で、入力値や基準パラメータを変えると再計算される。
                     ランクS/Aの行は条件付き書式で黄色ハイライト。
  スコアリング基準 : 配点定義・エリア点・ランク閾値のパラメータ表（数式の参照元）と前提。
  実績DB           : 接続検討回答の実績19件(2026-08-06更新)と分析メモ。
                     エリア＋変電所名の一致で実績メモ列に自動表示。

スコア構成（系統スコア80点。土地確定後に近接性S7=20点を加えて100点満点）:
  S1 空容量(上位系考慮) 35 / S2 N-1電制 10 / S3 潮流余裕 15 / S4 出力制御 5 / S5 配変 10 / S6 エリア 5
ゲート: 空容量(上位系考慮)≦0 かつ N-1電制≠可 → 除外（上位系増強リスク。実績: 喜多方=工期10年2ヶ月）
ランク: S:60点以上 / A:48-59 / B:36-47 / C:35以下 / データ無:空容量未公表(OSMのみ等)

配点・閾値・エリア点・実績DBの定義は substation_scoring.py に集約している
（Web版 substation.html 用データもそこを参照するため、定義の変更は同モジュールだけで済む）。

使い方:
  python scripts/score_substations.py --in 全国変電所リスト.xlsx --out 変電所スコアリング.xlsx
"""
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from substation_scoring import (
    AREA_POINTS, ANALYSIS_NOTES, NOTES, PARAM_ROWS, RESULTS_DB, SCORE_DEFS,
    S6_ACTUAL_STEPS, actual_months, s6_actual,
)

HEADER_FILL = PatternFill("solid", fgColor="FCE4D6")
PARAM_FILL = PatternFill("solid", fgColor="FFFF00")   # 変更可能なパラメータセル
RANK_FILL = PatternFill("solid", fgColor="FFF2A8")    # ランクS/Aハイライト（黄色）
GATE_FILL = PatternFill("solid", fgColor="E7E6E6")    # ゲート除外行（グレー）
HDR_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SMALL = Font(size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical="center")

BASE = "スコアリング基準"



def build_base_sheet(ws):
    """基準シートを作成し、数式が参照するパラメータセルの絶対参照を dict で返す。"""
    ws["A1"] = "変電所スコアリング基準（系統用蓄電池・高圧2MW前提）v1"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "作成: 2026-07-28 ｜ 接続検討の実回答12件の分析から逆算した配点で「接続検討を申し込む価値」を点数化する（実績DBは19件・2026-08-06更新。配点は19件ベースの分析でも変更なしを確認）"
    ws["A2"].font = SMALL

    r = 4
    ws.cell(r, 1, "変数・配点定義（実績12件で校正）").font = HDR_FONT
    r += 1
    for c, h in enumerate(["記号", "変数", "配点", "閾値・ルール", "根拠(実績)"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for row in SCORE_DEFS:
        r += 1
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v); cell.border = BORDER; cell.alignment = WRAP

    # 配点パラメータ（数式の参照元。黄色セルを変えるとスコアリングタブが再計算される）
    r += 2
    ws.cell(r, 1, "配点パラメータ（黄色セルを変更するとスコアが自動再計算されます）").font = HDR_FONT
    r += 1
    for c, h in enumerate(["項目", "閾値", "点数"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    params = {}
    for key, label, thr, pts in PARAM_ROWS:
        r += 1
        ws.cell(r, 1, label).border = BORDER
        tc = ws.cell(r, 2, thr); pc = ws.cell(r, 3, pts)
        for cell in (tc, pc):
            cell.border = BORDER; cell.alignment = CENTER
            if cell.value is not None:
                cell.fill = PARAM_FILL
        params[key] = {"thr": f"{BASE}!$B${r}", "pts": f"{BASE}!$C${r}"}

    # S6 エリア点表（VLOOKUP参照元）
    r += 2
    ws.cell(r, 1, "S6 エリア点（実績工期・出力制御環境からの仮置き。実績追加で更新）").font = HDR_FONT
    r += 1
    for c, h in enumerate(["エリア", "点数", "根拠(実績)"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    area_first = r + 1
    for area, pts, why in AREA_POINTS:
        r += 1
        ws.cell(r, 1, area).border = BORDER
        pc = ws.cell(r, 2, pts); pc.border = BORDER; pc.alignment = CENTER; pc.fill = PARAM_FILL
        ws.cell(r, 3, why).border = BORDER
    params["area_range"] = f"{BASE}!$A${area_first}:$B${r}"

    # ランク集計（スコアリングタブに連動）
    r += 2
    ws.cell(r, 1, "ランク集計（スコアリングタブに連動）").font = HDR_FONT
    r += 1
    for c, h in enumerate(["ランク", "件数", "定義"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    params["rank_count_row"] = r + 1  # 呼び出し側で件数のCOUNTIF数式を入れる

    return params


def add_rank_counts(ws, start_row, rank_col_range):
    defs = [
        ("S", "60点以上（最優先で周辺の土地を探す）"),
        ("A", "48〜59点（優先候補）"),
        ("B", "36〜47点"),
        ("C", "35点以下"),
        ("除外", "ゲート該当（空容量≦0 かつ N-1電制≠可 → 上位系増強リスク）"),
        ("データ無", "空容量未公表(OSMのみ等)。スコア比較不能"),
    ]
    for i, (rank, note) in enumerate(defs):
        r = start_row + i
        ws.cell(r, 1, rank).border = BORDER
        fc = ws.cell(r, 2, f'=COUNTIF({rank_col_range},"{rank}")')
        fc.border = BORDER; fc.alignment = CENTER
        ws.cell(r, 3, note).border = BORDER
    for col, w in [("A", 44), ("B", 14), ("C", 52), ("D", 52), ("E", 44)]:
        ws.column_dimensions[col].width = w


def build_results_sheet(ws):
    """実績DBシート。戻り値: (要約列レンジ, キー列レンジ)"""
    ws["A1"] = "接続検討回答 実績19件（2026-08-06更新。詳細: 接続検討回答_分析メモ_20260724.md＋追加実績メモ20260806）"
    ws["A1"].font = TITLE_FONT
    headers = ["エリア", "変電所名(リスト表記)", "案件", "出力", "負担金", "連系工期",
               "主な対策工事", "教訓", "要約(スコア表参照用)", "キー(自動)",
               "工期(月)", "S6点(実績)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for i, row in enumerate(RESULTS_DB):
        r = 3 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v); cell.border = BORDER; cell.alignment = WRAP
        kc = ws.cell(r, 10, f"=A{r}&TRIM(B{r})")
        kc.border = BORDER; kc.font = SMALL
        # 工期の表記が揺れるためPython側で月数に直し、そこからS6点を出しておく。
        # 同じ変電所に複数件あるときは最長を採る（工事規模で変わり、事前には決まらない）。
        mo = actual_months(row[5])
        ws.cell(r, 11, mo).border = BORDER
        ws.cell(r, 12, s6_actual(row[0], row[1])).border = BORDER
    last = 2 + len(RESULTS_DB)
    r = last + 2
    ws.cell(r, 1, "■分析メモ(2026-08-06追記・19件ベース)").font = HDR_FONT
    for i, note in enumerate(ANALYSIS_NOTES):
        ws.cell(r + 1 + i, 1, note).font = SMALL
    widths = [8, 20, 24, 12, 30, 30, 34, 30, 40, 18, 9, 11]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    return (f"実績DB!$I$3:$I${last}", f"実績DB!$J$3:$J${last}",
            f"実績DB!$L$3:$L${last}")


def find_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "No.":
            return row[0].row
    raise SystemExit("入力シートに 'No.' ヘッダー行が見つかりません")


def read_substations(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"シート '{sheet_name}' がありません。存在するシート: {wb.sheetnames}")
    ws = wb[sheet_name]
    hdr = find_header_row(ws)
    headers = [c.value or "" for c in ws[hdr]]

    def col(token):
        for i, h in enumerate(headers):
            if token in str(h).replace("\n", ""):
                return i
        raise SystemExit(f"ヘッダー '{token}' が見つかりません: {headers}")

    idx = {
        "no": col("No."), "area": col("エリア"), "name": col("変電所名"),
        "pref": col("都道府県"), "vmax": col("最大電圧"), "vsec": col("二次電圧"),
        "lat": col("緯度"), "lon": col("経度"), "acc": col("座標精度"),
        "cap": col("設備容量"), "opcap": col("運用容量"), "flow": col("予想潮流"),
        "avail": col("当該設備"), "avail_up": col("上位系考慮"),
        "n1": col("N-1電制"), "curtail": col("出力制御"), "addr": col("所在地"),
    }
    rows = []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if row[idx["no"]] is None and row[idx["name"]] is None:
            continue
        rows.append([row[idx[k]] for k in
                     ["no", "area", "name", "pref", "vmax", "vsec", "lat", "lon", "acc",
                      "cap", "opcap", "flow", "avail", "avail_up", "n1", "curtail", "addr"]])
    return rows


def build_scoring_sheet(ws, rows, p, memo_range, key_range, s6act_range):
    ws["A1"] = "全国変電所スコアリング（高圧2MW BESS）　黄色=ランクS/A ／ グレー=ゲート除外"
    ws["A1"].font = TITLE_FONT
    headers = ["No.", "エリア", "変電所名", "都道府県", "最大電圧kV", "二次電圧kV", "緯度", "経度",
               "座標精度", "設備容量MW", "運用容量MW", "予想潮流MW", "空容量_当該MW", "空容量_上位系MW",
               "N-1電制", "出力制御", "所在地(参考)", "ゲート", "S1空容量(35)", "S2 N-1(10)",
               "S3潮流(15)", "S4出制(5)", "S5配変(10)", "S6エリア(5)", "系統スコア(80)", "ランク", "実績メモ"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = HEADER_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER

    # 空容量: 上位系考慮(N列)を優先、無ければ当該設備(M列)で代用（基準シートの前提参照）
    eff = lambda r: f'IF($N{r}<>"",$N{r},$M{r})'

    for i, data in enumerate(rows):
        r = 3 + i
        for c, v in enumerate(data, 1):
            ws.cell(r, c, v)
        f = {
            # R ゲート: 空容量≦0 かつ N-1電制≠可 → 除外
            18: f'=IF(COUNT($M{r}:$N{r})=0,"",IF(AND({eff(r)}<=0,$O{r}<>"可"),"除外",""))',
            # S1 空容量(上位系考慮)
            19: f'=IF(COUNT($M{r}:$N{r})=0,0,IF({eff(r)}>={p["s1_t1"]["thr"]},{p["s1_t1"]["pts"]},'
                f'IF({eff(r)}>={p["s1_t2"]["thr"]},{p["s1_t2"]["pts"]},'
                f'IF({eff(r)}>0,{p["s1_t3"]["pts"]},0))))',
            # S2 N-1電制
            20: f'=IF($O{r}="可",{p["s2"]["pts"]},0)',
            # S3 潮流余裕 |予想潮流|/運用容量
            21: f'=IF(OR(COUNT($K{r}:$L{r})<2,$K{r}=0),0,'
                f'IF(ABS($L{r})/$K{r}<={p["s3_t1"]["thr"]},{p["s3_t1"]["pts"]},'
                f'IF(ABS($L{r})/$K{r}<={p["s3_t2"]["thr"]},{p["s3_t2"]["pts"]},'
                f'IF(ABS($L{r})/$K{r}<={p["s3_t3"]["thr"]},{p["s3_t3"]["pts"]},0))))',
            # S4 出力制御
            22: f'=IF($P{r}="なし",{p["s4"]["pts"]},0)',
            # S5 配電用変電所か(二次電圧)
            23: f'=IF($F{r}="",0,IF(ISNUMBER($F{r}),'
                f'IF($F{r}<={p["s5_1"]["thr"]},{p["s5_1"]["pts"]},{p["s5_2"]["pts"]}),0))',
            # S6 その変電所の実績工期があればそれを使い、無ければエリア点。
            # エリア点は工期を知らないときの代理値なので、実測があれば置き換える。
            24: f'=IFERROR(INDEX({s6act_range},MATCH($B{r}&$C{r},{key_range},0)),'
                f'IFERROR(VLOOKUP($B{r},{p["area_range"]},2,FALSE),0))',
            # 系統スコア(80)
            25: f'=SUM($S{r}:$X{r})',
            # ランク
            26: f'=IF($R{r}="除外","除外",IF(COUNT($M{r}:$N{r})=0,"データ無",'
                f'IF($Y{r}>={p["rank_s"]["thr"]},"S",IF($Y{r}>={p["rank_a"]["thr"]},"A",'
                f'IF($Y{r}>={p["rank_b"]["thr"]},"B","C")))))',
            # 実績メモ（実績DBのエリア＋変電所名キーと一致すれば要約を表示）
            27: f'=IFERROR(INDEX({memo_range},MATCH($B{r}&$C{r},{key_range},0)),"")',
        }
        for c, formula in f.items():
            ws.cell(r, c, formula)
        for c in range(18, 27):
            ws.cell(r, c).alignment = CENTER

    last = 2 + len(rows)
    rng = f"A3:AA{last}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=['$Z3="除外"'], fill=GATE_FILL, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['OR($Z3="S",$Z3="A")'], fill=RANK_FILL))
    ws.auto_filter.ref = f"A2:AA{last}"
    ws.freeze_panes = "D3"

    widths = [6, 8, 26, 11, 10, 10, 11, 11, 14, 10, 10, 10, 12, 13, 16, 9, 22,
              7, 11, 10, 10, 9, 10, 10, 12, 9, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for col in ("G", "H"):
        for r in range(3, last + 1):
            ws.cell(r, {"G": 7, "H": 8}[col]).number_format = "0.000000"
    return f"スコアリング!$Z$3:$Z${last}"


def main():
    ap = argparse.ArgumentParser(description="全国変電所リストxlsxから変電所スコアリングxlsxを生成")
    ap.add_argument("--in", dest="src", required=True, help="全国変電所リスト（66kV以上）のxlsxパス")
    ap.add_argument("--out", dest="out", required=True, help="出力xlsxパス")
    ap.add_argument("--list-sheet", default="全国変電所リスト", help="入力の変電所リストのシート名")
    args = ap.parse_args()

    rows = read_substations(args.src, args.list_sheet)
    print(f"読込: {len(rows)}行")

    wb = Workbook()
    ws_score = wb.active
    ws_score.title = "スコアリング"
    ws_base = wb.create_sheet(BASE)
    ws_db = wb.create_sheet("実績DB")

    params = build_base_sheet(ws_base)
    memo_range, key_range, s6act_range = build_results_sheet(ws_db)
    rank_range = build_scoring_sheet(ws_score, rows, params, memo_range, key_range,
                                     s6act_range)
    add_rank_counts(ws_base, params["rank_count_row"], rank_range)

    r = params["rank_count_row"] + 8
    ws_base.cell(r, 1, "前提・注意").font = HDR_FONT
    for i, note in enumerate(NOTES):
        ws_base.cell(r + 1 + i, 1, note).font = SMALL

    wb.save(args.out)
    print(f"出力: {args.out}")


if __name__ == "__main__":
    main()
