#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
候補地リスト（緯度経度つき）に、変電所スコアと土地のハザード判定を付ける。

用地の一覧を受け取り、1件ずつ次を出す。

  変電所 : 最寄りの変電所を距離順に評価し、系統スコア80 + 近接S7 20 = 総合100点
           （substation.html と同じ計算。substation_scoring.py を共用する）
  ハザード: 国土地理院「重ねるハザードマップ」のタイルを座標のピクセルで判定
           （洪水・津波・高潮・土砂災害3種）
  所在地  : 国土地理院リバースジオコーダで座標の市区町村を引き、入力の所在地と突き合わせる

【許認可17項目の自動判定は別】用途地域・市街化区域などは reinfolib のAPIキーが要る
（`reinfolib_judge.py`）。ここで出すのはキー不要で取れるハザードと所在地の確認まで。

【タイル判定の読み方】タイルが404＝その区画に該当区域のデータが無い＝非該当。
200なら座標のピクセルを見て、色が付いていれば該当。透明なら非該当。
ハザードマップは自治体の公表状況で整備範囲が違うため、非該当は「危険が無い」ではなく
「公表された区域に入っていない」を意味する。最終確認は自治体の窓口・図面で行うこと。

使い方:
  python scripts/screen_sites.py \
    --sites 候補地リスト.xlsx --subs "data/全国変電所リスト_66kV以上_座標追加.xlsx" \
    --out 候補地_スクリーニング.xlsx
"""
import argparse
import io
import math
import re
import urllib.request

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import substation_scoring as S
from score_substations import read_substations
from list_missing_coords import FIELDS

UA = {"User-Agent": "Mozilla/5.0"}
GSI_REV = ("https://mreversegeocoder.gsi.go.jp/reverse-geocoder/"
           "LonLatToAddress?lat={lat}&lon={lon}")

# 重ねるハザードマップのタイル。(表示名, URL雛形, ズーム)
HAZARDS = [
    ("洪水浸水(想定最大)", "https://disaportaldata.gsi.go.jp/raster/01_flood_l2_shinsuishin_data/{z}/{x}/{y}.png", 16),
    ("津波浸水想定", "https://disaportaldata.gsi.go.jp/raster/04_tsunami_newlegend_data/{z}/{x}/{y}.png", 16),
    ("高潮浸水想定", "https://disaportaldata.gsi.go.jp/raster/03_hightide_l2_shinsuishin_data/{z}/{x}/{y}.png", 16),
    ("土砂災害(土石流)", "https://disaportaldata.gsi.go.jp/raster/05_dosekiryukeikaikuiki/{z}/{x}/{y}.png", 16),
    ("土砂災害(急傾斜)", "https://disaportaldata.gsi.go.jp/raster/05_kyukeishakeikaikuiki/{z}/{x}/{y}.png", 16),
    ("土砂災害(地すべり)", "https://disaportaldata.gsi.go.jp/raster/05_jisuberikeikaikuiki/{z}/{x}/{y}.png", 16),
]

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
HIT_FILL = PatternFill("solid", fgColor="FEECEA")     # 赤=該当
OK_FILL = PatternFill("solid", fgColor="E8FAF3")      # 緑=ランクS/A
HEAD_FONT = Font(bold=True, size=10)


def parse_latlon(text):
    """「39.5907, 140.5725」形式から (lat, lon) を取る。読めなければ None。"""
    m = re.findall(r"-?\d+\.\d+", str(text or ""))
    if len(m) < 2:
        return None
    lat, lon = float(m[0]), float(m[1])
    if not (20 <= lat <= 46 and 122 <= lon <= 154):
        return None
    return lat, lon


def tile_xy(lat, lon, z):
    n = 2 ** z
    x = (lon + 180.0) / 360.0 * n
    y = (1.0 - math.log(math.tan(math.radians(lat))
                        + 1 / math.cos(math.radians(lat))) / math.pi) / 2.0 * n
    return x, y


def hazard_hit(lat, lon, url_tpl, z):
    """該当なら (True, 色), 非該当なら (False, None), 取得できなければ (None, 理由)。"""
    fx, fy = tile_xy(lat, lon, z)
    x, y = int(fx), int(fy)
    url = url_tpl.format(z=z, x=x, y=y)
    try:
        raw = urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=25).read()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return False, None          # その区画にデータが無い＝該当区域外
        return None, f"HTTP {e.code}"
    except Exception as e:
        return None, type(e).__name__
    from PIL import Image
    im = Image.open(io.BytesIO(raw)).convert("RGBA")
    px = int((fx - x) * im.width)
    py = int((fy - y) * im.height)
    px = min(max(px, 0), im.width - 1)
    py = min(max(py, 0), im.height - 1)
    r, g, b, a = im.getpixel((px, py))
    if a == 0:
        return False, None
    return True, f"#{r:02X}{g:02X}{b:02X}"


def reverse_geocode(lat, lon):
    try:
        import json
        d = json.load(urllib.request.urlopen(
            GSI_REV.format(lat=lat, lon=lon), timeout=20))["results"]
        return str(d.get("muniCd") or ""), str(d.get("lv01Nm") or "")
    except Exception:
        return "", ""


def haversine(a, b, c, d):
    r = 6371000.0
    p = math.radians
    x = (math.sin(p(c - a) / 2) ** 2
         + math.cos(p(a)) * math.cos(p(c)) * math.sin(p(d - b) / 2) ** 2)
    return 2 * r * math.asin(math.sqrt(x))


def s7_points(dist_m):
    for thr, pts in S.S7_STEPS:
        if dist_m <= thr:
            return pts
    return 0


def load_substations(path, sheet):
    rows = [dict(zip(FIELDS, r)) for r in read_substations(path, sheet)]
    out = []
    for r in rows:
        if not isinstance(r["lat"], (int, float)) or not isinstance(r["lon"], (int, float)):
            continue
        sc = S.score_row(r["area"], r["vsec"], r["opcap"], r["flow"],
                         r["avail"], r["avail_up"], r["n1"], r["curtail"], name=r["name"])
        r["sc"] = sc
        out.append(r)
    return out


def nearest(subs, lat, lon, topn, include_gate, radius_m):
    """候補地の周辺にある変電所を総合スコア順に返す。

    半径で先に絞る。絞らないと、遠方の系統スコアが高い変電所（S7=0でも80点満点中70点）が
    近傍の変電所を上回り、「最寄り」でも何でもないものが1位になる。
    S7は5km超で0点なので、それより広く採っても順位はほぼ系統スコアだけで決まる。
    """
    hits = []
    for r in subs:
        d = haversine(lat, lon, r["lat"], r["lon"])
        if d > radius_m:
            continue
        rank = r["sc"]["rank"]
        if not include_gate and rank in ("除外", "データ無"):
            continue
        s7 = s7_points(d)
        hits.append((r["sc"]["grid_score"] + s7, d, s7, r))
    hits.sort(key=lambda x: (-x[0], x[1]))
    return hits[:topn]


def read_sites(path):
    ws = load_workbook(path, data_only=True).worksheets[0]
    hdr = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or "") for c in row]
        if any("緯度" in v for v in vals):
            hdr = row[0].row
            heads = vals
            break
    if hdr is None:
        raise SystemExit("ヘッダー行（緯度経度を含む）が見つかりません")
    idx = {h.strip(): i for i, h in enumerate(heads) if h.strip()}
    sites = []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if all(v is None for v in row):
            continue
        rec = {k: row[i] for k, i in idx.items() if i < len(row)}
        ll = parse_latlon(rec.get("緯度経度"))
        if ll is None:
            continue
        rec["lat"], rec["lon"] = ll
        sites.append(rec)
    return sites, list(idx)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--sites", required=True, help="候補地リストのxlsx（緯度経度列が要る）")
    ap.add_argument("--subs", required=True, help="全国変電所リストのxlsx")
    ap.add_argument("--subs-sheet", default="全国変電所リスト")
    ap.add_argument("--out", required=True)
    ap.add_argument("--topn", type=int, default=3, help="1候補地あたり何件の変電所を出すか")
    ap.add_argument("--radius", type=float, default=10.0,
                    help="候補地から何kmまでの変電所を対象にするか（既定10km）。"
                         "S7は5km超で0点なので、広げても順位は系統スコアだけで決まる")
    ap.add_argument("--no-hazard", action="store_true", help="ハザード判定を省く（ネットワーク不要）")
    args = ap.parse_args()

    sites, site_cols = read_sites(args.sites)
    subs = load_substations(args.subs, args.subs_sheet)
    print(f"候補地 {len(sites)}件 / 座標のある変電所 {len(subs)}件")

    wb = Workbook()
    ws = wb.active
    ws.title = "スクリーニング"
    cols = ["No.", "所在（地番）", "土地面積", "用途", "地目", "緯度", "経度",
            "所在地(座標から)", "所在地の一致",
            "最寄り変電所", "エリア", "距離", "系統(80)", "S7(20)", "総合(100)", "ランク"]
    cols += [h for h, _, _ in HAZARDS] if not args.no_hazard else []
    ws.cell(1, 1, "候補地スクリーニング　緑=総合ランクS/A ／ 赤=ハザード該当").font = Font(bold=True, size=11)
    for c, h in enumerate(cols, 1):
        cell = ws.cell(2, c, h)
        cell.font = HEAD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    r = 3
    summary = []
    for s in sites:
        top = nearest(subs, s["lat"], s["lon"], args.topn, False, args.radius * 1000)
        haz = {}
        if not args.no_hazard:
            for name, tpl, z in HAZARDS:
                hit, info = hazard_hit(s["lat"], s["lon"], tpl, z)
                haz[name] = ("該当" if hit else "非該当") if hit is not None else f"取得不可({info})"
        muni, aza = reverse_geocode(s["lat"], s["lon"])
        addr = str(s.get("所在（地番）") or "")
        match = "—"
        if aza:
            match = "一致" if (aza and aza[:2] in addr) else "要確認"
        first = True
        for total, dist, s7, sub in (top or [(None, None, None, None)]):
            vals = [s.get("地図NO") if first else "", addr if first else "",
                    s.get("土地面積 ") or s.get("土地面積") if first else "",
                    s.get("用途") if first else "", s.get("地目") if first else "",
                    round(s["lat"], 6) if first else "", round(s["lon"], 6) if first else "",
                    (f"{muni} {aza}".strip() if first else ""), (match if first else "")]
            if sub is None:
                vals += [f"（半径{args.radius:g}km内に該当なし）", "", "", "", "", "", ""]
            else:
                vals += [sub["name"], sub["area"],
                         f"{dist/1000:.2f} km" if dist >= 1000 else f"{int(dist)} m",
                         sub["sc"]["grid_score"], s7, total, sub["sc"]["rank"]]
            if not args.no_hazard:
                vals += [haz[n] if first else "" for n, _, _ in HAZARDS]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                if not args.no_hazard and c > len(cols) - len(HAZARDS) and v == "該当":
                    cell.fill = HIT_FILL
                if c <= 2 and sub is not None and total is not None and total >= 60:
                    cell.fill = OK_FILL
            r += 1
            first = False
        if top:
            t = top[0]
            summary.append((s.get("地図NO"), addr, t[3]["name"], t[1], t[0], t[3]["sc"]["rank"],
                            [n for n in haz if haz.get(n) == "該当"]))
        else:
            summary.append((s.get("地図NO"), addr, None, None, None, None,
                            [n for n in haz if haz.get(n) == "該当"]))

    widths = [6, 34, 10, 8, 12, 10, 11, 18, 9, 18, 7, 9, 8, 7, 8, 7] + [13] * len(HAZARDS)
    for c, w in enumerate(widths[:len(cols)], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{r-1}"

    note = wb.create_sheet("読み方・出典")
    lines = [
        ("変電所スコア", "系統スコア80点（S1空容量35/S2 N-1電制10/S3潮流15/S4出力制御5/S5配変10/S6エリア・実績5）"),
        ("", "＋近接スコアS7 20点＝総合100点。substation.html と同じ計算。"),
        ("", "ゲート除外・データ無の変電所は候補から外している（申し込み対象にならないため）。"),
        ("", "総合ランクは S:75以上 / A:60 / B:45。"),
        ("", ""),
        ("ハザード判定", "国土地理院「重ねるハザードマップ」のタイルを座標のピクセルで判定している。"),
        ("", "タイルが無い区画＝その災害の公表区域が無い＝非該当。"),
        ("", "**非該当は「危険が無い」ではなく「公表された区域に入っていない」**。"),
        ("", "整備範囲は自治体によって違うため、最終確認は自治体の窓口・図面で行うこと。"),
        ("", ""),
        ("所在地の一致", "座標を国土地理院リバースジオコーダで引き、入力の地番と突き合わせた参考値。"),
        ("", "「要確認」は座標か地番のどちらかがずれている可能性を示す。"),
        ("", ""),
        ("この表に無いもの", "用途地域・市街化区域/調整区域・液状化・自然公園などの許認可17項目は"),
        ("", "reinfolib のAPIキーが要るため含めていない（scripts/reinfolib_judge.py）。"),
        ("", "座標が確定しているので、キーがあればそのまま自動判定できる。"),
    ]
    for i, (k, v) in enumerate(lines):
        note.cell(i + 1, 1, k).font = Font(bold=True, size=10)
        note.cell(i + 1, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 92

    wb.save(args.out)
    print(f"出力: {args.out}")
    print()
    print(f"{'No':>3s} {'最寄り変電所':14s}{'距離':>9s}{'総合':>5s} {'ランク':4s} ハザード該当")
    for no, addr, nm, dist, total, rank, hz in summary:
        if nm is None:
            print(f"{str(no):>3s} （該当なし）")
            continue
        d = f"{dist/1000:.2f}km" if dist >= 1000 else f"{int(dist)}m"
        print(f"{str(no):>3s} {nm[:13]:14s}{d:>9s}{total:>5d} {rank:4s} {'/'.join(hz) if hz else '—'}")


if __name__ == "__main__":
    main()
