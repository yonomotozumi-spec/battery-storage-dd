#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
変電所の諸元をExcelにまとめる（PX収支シミュレーターへの受け渡しを兼ねる）。

`score_substations.py` が出すのは採点結果のワークブックで、数式と配点が主役になる。
こちらは**設備の諸元と、PX収支シミュレーターの入力に必要な値**を1行1変電所で並べた
参照用の表。候補地の検討時にフィルタして使う。

PX収支シミュレーター（v1.47）は変電所側から次を必要とする。
  エリア（プルダウン：北海道/東北/東京/中部/北陸/関西/中国/四国/九州）
  電圧区分（特高／高圧）→ 託送料金区分「エリア-区分」が決まる
  連系出力kW（空容量が上限の目安）

そこで各行に **電圧区分・託送料金区分・基本料金・従量料金・損失率・発電側課金** を付ける。
電圧区分は二次電圧で判定する（≦7kV＝配電用変電所なら高圧連系が可能）。S5の判定と同じ基準。

【沖縄はシミュレーター未対応】エリアのプルダウンに沖縄が無く、託送料金マスタにも
沖縄の行が無い。沖縄の変電所は諸元だけ載せ、PX入力欄は「PXシミュレーター未対応」とする。

使い方:
  python scripts/build_substation_specs.py \
    --in "data/全国変電所リスト_66kV以上_座標追加.xlsx" \
    --out "docs/変電所諸元_全国66kV以上.xlsx"
"""
import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import substation_scoring as S
from list_missing_coords import FIELDS
from score_substations import read_substations

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
PX_FILL = PatternFill("solid", fgColor="FFF2CC")     # PXシミュレーターへ渡す列
SA_FILL = PatternFill("solid", fgColor="E8FAF3")     # ランクS/A
HEAD_FONT = Font(bold=True, size=10)

# 託送料金（税込）。出典: PX収支シミュレーター v1.47「参考2.託送料金系マスタ」
# （2025/8/7公表・2025/10適用）。エリア → 区分 → (基本[円/kW月], 従量[円/kWh], 損失率)
TARIFF = {
    "北海道": {"特高": (515.9, 1.02, 0.020), "高圧": (842.6, 2.28, 0.047), "発電側課金": 110.0},
    "東北":   {"特高": (462.0, 0.97, 0.019), "高圧": (728.2, 2.15, 0.052), "発電側課金": 93.04},
    "東京":   {"特高": (423.39, 0.91, 0.013), "高圧": (653.87, 1.84, 0.037), "発電側課金": 87.01},
    "中部":   {"特高": (357.5, 0.88, 0.025), "高圧": (467.5, 2.21, 0.038), "発電側課金": 80.42},
    "北陸":   {"特高": (572.0, 0.85, 0.013), "高圧": (748.0, 1.76, 0.034), "発電側課金": 93.47},
    "関西":   {"特高": (440.0, 0.84, 0.029), "高圧": (663.3, 2.29, 0.043), "発電側課金": 97.98},
    "中国":   {"特高": (495.0, 0.86, 0.021), "高圧": (748.0, 2.16, 0.044), "発電側課金": 92.4},
    "四国":   {"特高": (511.5, 1.03, 0.023), "高圧": (759.0, 2.35, 0.048), "発電側課金": 101.2},
    "九州":   {"特高": (394.9, 0.83, 0.019), "高圧": (629.2, 2.06, 0.043), "発電側課金": 79.53},
}
NO_PX = "PXシミュレーター未対応"
NO_VSEC = "二次電圧が未公表で判定不可"

COLUMNS = [
    ("No.", 7), ("エリア", 8), ("都道府県", 10), ("変電所名", 22),
    ("最大電圧\nkV", 8), ("二次電圧\nkV", 8),
    ("空容量\n上位系MW", 10), ("空容量\n当該MW", 9), ("採用空容量\nMW", 10),
    ("運用容量\nMW", 9), ("予想潮流\nMW", 9), ("潮流率", 8),
    ("N-1電制", 14), ("出力制御", 10),
    ("緯度", 11), ("経度", 11), ("座標精度", 20),
    ("S1", 5), ("S2", 5), ("S3", 5), ("S4", 5), ("S5", 5), ("S6", 5),
    ("系統スコア\n80点", 10), ("ランク", 7), ("ゲート", 8),
    ("高圧連系", 9),
    ("PX:電圧区分", 11), ("PX:託送料金区分", 15),
    ("PX:基本料金\n円/kW月", 12), ("PX:従量料金\n円/kWh", 12),
    ("PX:損失率", 10), ("PX:発電側課金\n円/kW月", 13),
    ("地図", 8),
]
PX_FROM = 27   # 「PX:電圧区分」の列番号（1始まり）


def voltage_class(vsec):
    """二次電圧から連系の電圧区分を決める。≦7kVなら高圧連系が可能（S5と同じ基準）。"""
    v = S._num(vsec)
    if v is None:
        return "", ""
    if v <= S.P["s5_1"]["thr"]:
        return "高圧", "可"
    return "特高", "不可（二次電圧が高圧より上）"


def map_link(name, pref, area):
    import urllib.parse
    q = " ".join(x for x in (pref or area, str(name)) if x)
    return "https://www.google.com/maps/search/" + urllib.parse.quote(q)


def build_rows(src, sheet):
    rows = [dict(zip(FIELDS, r)) for r in read_substations(src, sheet)]
    out = []
    for r in rows:
        sc = S.score_row(r["area"], r["vsec"], r["opcap"], r["flow"],
                         r["avail"], r["avail_up"], r["n1"], r["curtail"])
        eff = S.effective_capacity(r["avail"], r["avail_up"])
        cap, fl = S._num(r["opcap"]), S._num(r["flow"])
        ratio = (abs(fl) / cap) if (cap and fl is not None and cap != 0) else None
        vclass, hv = voltage_class(r["vsec"])
        t = TARIFF.get(r["area"])
        if t is None:
            px = [NO_PX, "", "", "", "", ""]
        elif not vclass:
            # 二次電圧が未公表だと高圧／特高を決められない（東京エリアに多い）。
            # 空欄にすると「該当なし」と区別が付かないので理由を書く。
            px = [NO_VSEC, "", "", "", "", ""]
        else:
            basic, vol, loss = t[vclass]
            px = [vclass, f"{r['area']}-{vclass}", basic, vol, loss, t["発電側課金"]]
        out.append([
            r["no"], r["area"], r["pref"], r["name"],
            S._num(r["vmax"]), S._num(r["vsec"]),
            S._num(r["avail_up"]), S._num(r["avail"]), eff,
            cap, fl, round(ratio, 3) if ratio is not None else None,
            r["n1"], r["curtail"],
            r["lat"], r["lon"], r["acc"],
            sc["s1"], sc["s2"], sc["s3"], sc["s4"], sc["s5"], sc["s6"],
            sc["grid_score"], sc["rank"], sc["gate"],
            hv,
        ] + px + [map_link(r["name"], r["pref"], r["area"])])
    return out


def write_sheet(ws, title, data):
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=11)
    for c, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(2, c)
        cell.value = name
        cell.font = HEAD_FONT
        cell.fill = PX_FILL if c >= PX_FROM else HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width
    for i, row in enumerate(data):
        r = i + 3
        for c, v in enumerate(row, start=1):
            cell = ws.cell(r, c)
            if c == len(COLUMNS):
                cell.value = "地図"
                cell.hyperlink = v
                cell.style = "Hyperlink"
            else:
                cell.value = v
            if row[24] in ("S", "A") and c <= 4:
                cell.fill = SA_FILL
    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(data) + 2}"


def write_reference(ws):
    ws.cell(1, 1).value = "PX収支シミュレーターへの受け渡し"
    ws.cell(1, 1).font = Font(bold=True, size=11)
    lines = [
        ("この表の使い方", ""),
        ("", "諸元シートで候補の変電所を絞り、黄色の『PX:』列をシミュレーターの"
             "「1.各条件インプット」へ写す。"),
        ("", "エリアと電圧区分はプルダウンの選択肢と同じ表記にしてある。"),
        ("", "連系出力(kW)は採用空容量が上限の目安。高圧2MW前提なら1,990kW程度。"),
        ("", ""),
        ("電圧区分の判定", "二次電圧 ≦ 7kV（配電用変電所）→ 高圧連系が可能。"
                          "それ以外は特高。スコアのS5と同じ基準。"),
        ("", ""),
        ("沖縄について", "PX収支シミュレーター v1.47 はエリアのプルダウンに沖縄が無く、"
                        "託送料金マスタにも沖縄の行が無い。"),
        ("", "沖縄の変電所は諸元だけ載せ、PX入力欄は「" + NO_PX + "」としている。"),
        ("", ""),
        ("二次電圧が無い行", "二次電圧が未公表だと高圧／特高を決められないため、PX入力欄は"
                            "「" + NO_VSEC + "」としている。東京エリアに多い（1,644件中1,611件）。"),
        ("", "この場合でも諸元・座標・系統スコアは有効。電圧区分は接続検討で確認すること。"),
        ("", ""),
        ("託送料金の出典", "PX収支シミュレーター v1.47「参考2.託送料金系マスタ」"
                          "（2025/8/7公表・2025/10適用）。税込。"),
        ("", "基本料金は円/kW月、従量料金は円/kWh、発電側課金は円/kW月。"),
        ("", ""),
        ("注意", "空容量・予想潮流は送配電事業者の公表値で、時点が各社で異なる。"
                "実際の連系可否は接続検討の回答による。"),
        ("", "座標精度が『Google Places(名称一致)』の行は最大1.2kmの誤差がある"
             "（既知の中部496件で測定。中央値6m）。"),
        ("", "座標精度が『推定』の行は市町村レベルの推定値で、距離の判定には使えない。"),
    ]
    for i, (k, v) in enumerate(lines):
        ws.cell(i + 3, 1).value = k
        ws.cell(i + 3, 1).font = Font(bold=True, size=10)
        ws.cell(i + 3, 2).value = v
        ws.cell(i + 3, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 86

    base = len(lines) + 5
    ws.cell(base, 1).value = "託送料金マスタ（エリア別）"
    ws.cell(base, 1).font = Font(bold=True, size=11)
    head = ["エリア", "区分", "託送料金区分", "基本料金\n円/kW月",
            "従量料金\n円/kWh", "損失率", "発電側課金\n円/kW月"]
    for c, h in enumerate(head, start=1):
        cell = ws.cell(base + 1, c)
        cell.value = h
        cell.font = HEAD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True)
    r = base + 2
    for area, t in TARIFF.items():
        for kind in ("特高", "高圧"):
            basic, vol, loss = t[kind]
            for c, v in enumerate([area, kind, f"{area}-{kind}", basic, vol, loss,
                                   t["発電側課金"]], start=1):
                ws.cell(r, c).value = v
            r += 1
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 14


def main():
    ap = argparse.ArgumentParser(description="変電所の諸元をExcelにまとめる")
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--list-sheet", default="全国変電所リスト")
    args = ap.parse_args()

    data = build_rows(args.src, args.list_sheet)
    # S/Aシートは検討する順に並べる（系統スコアの降順。同点はエリア→名称で安定させる）
    sa = sorted((r for r in data if r[24] in ("S", "A")),
                key=lambda r: (-r[23], str(r[1]), str(r[3])))

    wb = Workbook()
    write_sheet(wb.active, "変電所諸元（全国66kV以上）　黄色=PX収支シミュレーターへ渡す列 ／ 緑=ランクS/A", data)
    wb.active.title = "諸元"
    write_sheet(wb.create_sheet("ランクS_A"),
                "ランクS/Aのみ（系統スコア48点以上・ゲート除外を除く）", sa)
    write_reference(wb.create_sheet("PX入力と出典"))
    wb.save(args.out)

    coords = sum(1 for r in data if isinstance(r[14], (int, float)))
    print(f"諸元 {len(data)}件（うちランクS/A {len(sa)}件 / 座標あり {coords}件）")
    print(f"出力: {args.out}")


if __name__ == "__main__":
    main()
