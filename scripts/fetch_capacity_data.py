#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
各社が公表する「予想潮流等一覧表」から、空容量データが欠測の行を埋める。

元リストは6,071件中1,507件で空容量が両列とも未公表で、ランクが「データ無」になる。
S1（35点）とゲート判定が効かないため実質的に評価できていない。北海道に至っては
385件すべてが欠測している。

各社はOCCTOの情報公表ルールに沿って**同一様式のCSV**（変圧器の予想潮流等一覧表）を
出しており、スコアリングに必要な項目が揃っている。

  変電所No, 変電所名, 電圧(一次)(kV), 電圧(二次)(kV), 台数, 設備容量(MW),
  運用容量値(MW), 運用容量制約要因, 予想潮流(MW),
  空容量(当該設備)(MW), 空容量(上位系等考慮)(MW),
  N-1電制適用可否, N-1電制適用可能量(MW), 平常時出力制御の可能性

【1変電所に複数行ある】バンク（電圧階級）ごとに1行で、値が違う。

  北芽室変電所  66kV  : 空容量 ―    N-1電制「可」
  北芽室変電所  6.6kV : 空容量 7MW  N-1電制「不可 ♯３」

高圧2MWの連系先は6.6kVバンクなので、採る行を間違えると静かに誤ったスコアになる。
**二次電圧で突き合わせ**、決められない場合は埋めない。

【既存の値は上書きしない】各社の公表時点はリストの値と異なる。欠測のセルだけを埋める。

使い方:
  python scripts/fetch_capacity_data.py fetch --cache-dir capacity_cache
  python scripts/fetch_capacity_data.py fill \
    --in "data/全国変電所リスト_66kV以上_座標追加.xlsx" \
    --out "data/全国変電所リスト_66kV以上_座標追加.xlsx" \
    --cache-dir capacity_cache --report capacity_report.csv
"""
import argparse
import csv
import io
import os
import re
import urllib.parse
import urllib.request
import zipfile

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from fetch_substation_coords import loose_name, normalize_name
from score_substations import find_header_row

UA = {"User-Agent": "Mozilla/5.0"}
FILL = PatternFill("solid", fgColor="E2EFDA")   # 緑=公表資料から補完
SRC_LABEL = "予想潮流等一覧表"

# エリアごとの公表ページと、変圧器の一覧表へのリンクを拾う正規表現。
# URLに公表年月が入って変わるため、固定せずページから拾い直す。
SOURCES = {
    "北海道": {
        "page": "https://www.hepco.co.jp/network/con_service/public_document/bid_info.html",
        "pat": r"zip/sys_capa[^\"']*\.zip",
    },
    "北陸": {
        "page": "https://www.rikuden.co.jp/nw_notification/U_154seiyaku.html",
        "pat": r"[^\"']*sys_capa[^\"']*_tr_[^\"']*\.csv",
    },
    "四国": {
        "page": "https://www.yonden.co.jp/nw/line_access/local/index.html",
        "pat": r"[^\"']*sys_capa[^\"']*_tr_[^\"']*\.csv",
    },
    "関西": {
        "page": "https://www.kansai-td.co.jp/consignment/disclosure/distribution-equipment/index.html",
        "pat": r"[^\"']*154kv_(?:more|less)_trans\.csv",
    },
}

# 列名は社ごとに揺れる（括弧の全半角、語順、「空き容量」表記）。含まれる語の組で判定する。
#   北海道/四国 : 空容量(当該設備)(MW) / 空容量(上位系等考慮)(MW)
#   関西        : 当該設備 空容量（MW） / 上位系等考慮 空容量（MW）
#   北陸        : 空き容量[MW] のみ（単一値）→ 当該設備として扱う
#                 ※READMEの方針どおり、単一値のみ公表のエリアは「当該設備」で代用する
def _has(h, *words):
    t = str(h).replace(" ", "").replace("　", "")
    return all(w in t for w in words)


COL_MATCH = {
    "name": lambda h: _has(h, "変電所名"),
    "vsec": lambda h: _has(h, "二次") and _has(h, "電圧"),
    "opcap": lambda h: _has(h, "運用容量値"),
    "flow": lambda h: _has(h, "予想潮流"),
    "avail": lambda h: (_has(h, "当該設備") and _has(h, "空容量")) or _has(h, "空き容量"),
    "avail_up": lambda h: _has(h, "上位系") and (_has(h, "空容量") or _has(h, "空き容量")),
    "n1": lambda h: _has(h, "N-1電制適用可否"),
    "curtail": lambda h: _has(h, "出力制御"),
}


def decode(raw):
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("cp932", errors="replace")


def num(v):
    """「―」「－」「-」などの欠測記号を None にする。"""
    s = str(v or "").replace(",", "").replace("　", "").strip()
    if s in ("", "―", "－", "-", "ー", "‐"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def text(v):
    s = str(v or "").replace("　", " ").strip()
    return "" if s in ("―", "－", "-", "ー", "‐") else s


def fetch(args):
    for area, cfg in SOURCES.items():
        d = os.path.join(args.cache_dir, area)
        os.makedirs(d, exist_ok=True)
        try:
            html = decode(urllib.request.urlopen(
                urllib.request.Request(cfg["page"], headers=UA), timeout=60).read())
        except Exception as e:
            print(f"{area}: 公表ページを取得できません {str(e)[:50]}")
            continue
        urls = sorted({urllib.parse.urljoin(cfg["page"], u)
                       for u in re.findall(cfg["pat"], html)})
        got = 0
        for u in urls:
            fn = os.path.join(d, u.rsplit("/", 1)[1])
            if os.path.exists(fn) and os.path.getsize(fn) > 200:
                got += 1
                continue
            try:
                data = urllib.request.urlopen(
                    urllib.request.Request(u, headers=UA), timeout=120).read()
            except Exception as e:
                print(f"  NG {u.rsplit('/', 1)[1]} {str(e)[:40]}")
                continue
            with open(fn, "wb") as f:
                f.write(data)
            got += 1
        print(f"{area}: {got}/{len(urls)}件を {d} に保存")


def read_tables(cache_dir):
    """キャッシュのCSV/ZIPを読み、エリア → 照合キー → バンク行のリスト を返す。"""
    out = {}
    for area in sorted(os.listdir(cache_dir)):
        d = os.path.join(cache_dir, area)
        if not os.path.isdir(d):
            continue
        texts = []
        for fn in sorted(os.listdir(d)):
            p = os.path.join(d, fn)
            if fn.endswith(".zip"):
                try:
                    z = zipfile.ZipFile(p)
                except zipfile.BadZipFile:
                    continue
                for n in z.namelist():
                    if re.search(r"_(?:Tr|tr|trans)_?", n) and n.endswith(".csv"):
                        texts.append(decode(z.read(n)))
            elif fn.endswith(".csv"):
                texts.append(decode(open(p, "rb").read()))

        idx = {}
        for txt in texts:
            hdr = None
            for row in csv.reader(io.StringIO(txt)):
                if not row:
                    continue
                joined = "".join(row)
                if hdr is None:
                    if "変電所名" in joined and ("空容量" in joined or "運用容量" in joined):
                        hdr = row
                    continue
                rec = dict(zip(hdr, row))

                def pick(key):
                    match = COL_MATCH[key]
                    for h, v in rec.items():
                        if h and match(h):
                            return v
                    return None

                name = text(pick("name"))
                if not name:
                    continue
                k = normalize_name(name)
                if not k:
                    # 「木代」「中能登」のように種別が付かない社があり、その場合は
                    # normalize_name が空にならず緩いキーと同じ値になる
                    k = loose_name(name)
                if not k:
                    continue
                idx.setdefault(k, []).append({
                    "name": name, "vsec": num(pick("vsec")),
                    "opcap": num(pick("opcap")), "flow": num(pick("flow")),
                    "avail": num(pick("avail")), "avail_up": num(pick("avail_up")),
                    "n1": text(pick("n1")), "curtail": text(pick("curtail")),
                })
        if idx:
            out[area] = idx
    return out


def choose_bank(banks, vsec):
    """採用するバンクを決める。(行, 理由) を返す。決められないときは (None, 理由)。"""
    withcap = [b for b in banks if b["avail"] is not None or b["avail_up"] is not None]
    if not withcap:
        return None, "一覧表に空容量の数値が無い"
    if vsec is not None:
        same = [b for b in withcap if b["vsec"] is not None and abs(b["vsec"] - vsec) < 0.05]
        if len(same) == 1:
            return same[0], f"二次電圧{vsec}kVで一致"
        if len(same) > 1:
            return None, f"二次電圧{vsec}kVのバンクが{len(same)}件あり特定できない"
        return None, f"二次電圧{vsec}kVに一致するバンクが一覧表に無い"
    # リスト側に二次電圧が無い場合は、空容量を持つバンクが1件のときだけ採る
    if len(withcap) == 1:
        return withcap[0], "二次電圧は不明だが空容量を持つバンクが1件"
    return None, f"二次電圧が不明で、空容量を持つバンクが{len(withcap)}件あり特定できない"


def fill(args):
    tables = read_tables(args.cache_dir)
    if not tables:
        raise SystemExit(f"{args.cache_dir} に一覧表がありません。先に fetch を実行してください")
    print("読み込んだエリア:", {a: len(v) for a, v in tables.items()})

    wb = load_workbook(args.src)
    ws = wb[args.list_sheet]
    hdr = find_header_row(ws)
    heads = [str(c.value or "").replace("\n", "") for c in ws[hdr]]

    def col(token):
        for i, h in enumerate(heads):
            if token in h:
                return i + 1
        raise SystemExit(f"ヘッダー '{token}' が見つかりません")

    C = {"name": col("変電所名"), "area": col("エリア"), "vsec": col("二次電圧"),
         "opcap": col("運用容量"), "flow": col("予想潮流"),
         "avail": col("当該設備"), "avail_up": col("上位系考慮"),
         "n1": col("N-1電制"), "curtail": col("出力制御")}

    report, filled, skipped = [], 0, 0
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, C["name"]).value
        if not name:
            continue
        area = ws.cell(r, C["area"]).value
        cur = num(ws.cell(r, C["avail"]).value)
        up = num(ws.cell(r, C["avail_up"]).value)
        if cur is not None or up is not None:
            continue                      # 既に空容量がある行は触らない
        idx = tables.get(area)
        if not idx:
            continue                      # そのエリアの一覧表を持っていない
        banks = idx.get(normalize_name(name))
        if not banks and str(name).strip().endswith("変電所"):
            # 関西・北陸は一覧表の名称に種別が付かない（「木代」「中能登」）。
            # 種別なしで索引された項目だけが緩いキーで引けるので、
            # 「芳賀変電所」と「芳賀開閉所」のような取り違えは起きない。
            banks = idx.get(loose_name(name))
        if not banks:
            skipped += 1
            report.append([area, name, "", "不採用", "一覧表に同名が無い"])
            continue
        bank, why = choose_bank(banks, num(ws.cell(r, C["vsec"]).value))
        if bank is None:
            skipped += 1
            report.append([area, name, "", "不採用", why])
            continue
        # 欠測しているセルだけを埋める
        wrote = []
        for key in ("avail", "avail_up", "opcap", "flow", "n1", "curtail"):
            v = bank[key]
            if v is None or v == "":
                continue
            cell = ws.cell(r, C[key])
            if num(cell.value) is not None or text(cell.value):
                continue
            cell.value = v
            cell.fill = FILL
            wrote.append(key)
        if not wrote:
            skipped += 1
            report.append([area, name, "", "不採用", "埋める欄が無い（既に値がある）"])
            continue
        filled += 1
        report.append([area, name, bank["name"], "採用",
                       f"{why} / 記入: {','.join(wrote)} / "
                       f"空容量 当該{bank['avail']} 上位系{bank['avail_up']}"])

    wb.save(args.out)
    print(f"空容量が欠測の行: 補完 {filled}件 / 見送り {skipped}件")
    print(f"出力: {args.out}")
    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["エリア", "変電所名", "一覧表の名称", "判定", "理由"])
            w.writerows(report)
        print(f"レポート: {args.report}")
    if filled:
        print("※ 反映後は score_substations.py と build_substation_web_data.py の再実行が必要です。")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    f1 = sub.add_parser("fetch", help="各社の予想潮流等一覧表を取得する")
    f1.add_argument("--cache-dir", default="capacity_cache")
    f2 = sub.add_parser("fill", help="欠測の空容量だけを埋める")
    f2.add_argument("--in", dest="src", required=True)
    f2.add_argument("--out", required=True)
    f2.add_argument("--cache-dir", default="capacity_cache")
    f2.add_argument("--list-sheet", default="全国変電所リスト")
    f2.add_argument("--report")
    args = ap.parse_args()
    fetch(args) if args.cmd == "fetch" else fill(args)


if __name__ == "__main__":
    main()
