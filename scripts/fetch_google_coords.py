#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Google Maps Places API (New) の Text Search から変電所座標を取得する。

各社の系統情報公表資料は中部電力PGを除き模式図のPDFで座標を持たない（README参照）。
一方 Google は日本の変電所をPOIとして登録しており、都道府県＋変電所名で引ける。

【誤マッチが出るので必ず名称で検証する】Googleは該当が無くても「該当なし」を返さず、
近隣の別変電所を返す。実測した誤りの例:

  北大出変電所 → 新北信変電所（92.9km）   上地変電所 → 下地変電所（24.6km）
  笹部変電所   → 笹賀変電所（5.1km）      関ヶ原変電所 → 関ヶ原開閉所（5.3km）

都道府県では弾けない（クエリに県を含めるため住所は496/496が県一致した）。
fetch_substation_coords.py の normalize_name で**Google側の表示名も正規化**して
一致したものだけを採用する。事業者名プレフィックスは吸収され、施設種別は区別される。

【名称一致だけでは足りない】名称が完全に一致していても、別地方の同名変電所を掴むことがある。
実測では474件中26件（5.5%）がこれだった。

  渋谷変電所（関西）→ 東京都   長岡変電所（関西）→ 新潟県   海南変電所（徳島県）→ 和歌山県

そのため採用前に**地理院リバースジオコーダで座標の所在地を引き直し**、行の都道府県と
一致するかを必ず確認する。都道府県が空の行（関西・九州に多い）は、都道府県→エリアの
対応をリスト自身から作ってエリアの一致で確認する。誤りは都道府県が既知の行で1.2%、
空の行で15.5%あり、この確認なしでは使えない。

【中部496件（公表資料の正解）での実測】
  採用 444件: 誤差の中央値 6m / 最大 1,246m / 500m以内 439件 / 5km超 0件 / 同一座標の重複 0件
  除外  52件: 誤差の中央値 2,596m / 最大 92.9km（うち19件が5km超）
  ※ 除外のうち14件は実際には正しく、安全側に倒した分の取りこぼし

使い方:
  export GOOGLE_MAPS_API_KEY=...        # キーはリポジトリに置かない

  # ① 精度を測る（座標が既知の行で誤差を出す。新しいエリアに使う前に必ず実行する）
  python scripts/fetch_google_coords.py validate \
    --in "data/全国変電所リスト_66kV以上_座標追加.xlsx" --acc "公表資料(手動)"

  # ② 優先リストの空欄を埋める（所在地の確認まで行う）
  python scripts/fetch_google_coords.py fill \
    --in "data/全国変電所リスト_66kV以上_座標追加.xlsx" \
    --list "docs/座標未取得_優先リスト.xlsx" --report google_fill_report.csv
"""
import argparse
import collections
import concurrent.futures as cf
import csv
import json
import math
import os
import statistics
import urllib.request

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from fetch_substation_coords import normalize_name, read_rows

URL = "https://places.googleapis.com/v1/places:searchText"
GSI_REV = ("https://mreversegeocoder.gsi.go.jp/reverse-geocoder/"
           "LonLatToAddress?lat={lat}&lon={lon}")
FIELDS = "places.displayName,places.location,places.formattedAddress"
SOURCE_LABEL = "Google Places(名称一致)"
FILLED_FILL = PatternFill("solid", fgColor="FFF2CC")
WORKERS = 8


def search(name, pref, key):
    """(lat, lon, 表示名, 住所) or None。名称が一致しない候補は捨てる。"""
    body = json.dumps({"textQuery": f"{pref or ''} {name}".strip(),
                       "languageCode": "ja", "regionCode": "JP"}).encode()
    req = urllib.request.Request(URL, data=body, headers={
        "Content-Type": "application/json", "X-Goog-Api-Key": key,
        "X-Goog-FieldMask": FIELDS})
    try:
        res = json.load(urllib.request.urlopen(req, timeout=30))
    except Exception as e:
        return None, f"照会に失敗: {str(e)[:60]}"
    places = res.get("places") or []
    if not places:
        return None, "該当なし"
    want = normalize_name(name)
    if not want:
        return None, "正規化して固有名が残らない名前"
    for p in places:
        got = p.get("displayName", {}).get("text", "")
        if normalize_name(got) == want:
            return (p["location"]["latitude"], p["location"]["longitude"],
                    got, p.get("formattedAddress", "")), "採用"
    return None, f"名称が一致しない（Google側は「{places[0].get('displayName',{}).get('text','')}」）"


def pref_to_area(rows):
    """都道府県→エリアの対応をリスト自身から作る（固定の地理データを持たない）。"""
    tally = {}
    for r in rows:
        if r["pref"] and r["area"]:
            tally.setdefault(r["pref"], collections.Counter())[r["area"]] += 1
    return {p: c.most_common(1)[0][0] for p, c in tally.items()}


def muni_code_to_pref(rows):
    """全国地方公共団体コードの上2桁→都道府県名。リストにある都道府県から作る。"""
    # 既知座標を1件ずつ引くのは高くつくので、コード表は都道府県名の並び順で持つ
    order = ["北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
             "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
             "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
             "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
             "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
             "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
             "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"]
    return {f"{i:02d}": p for i, p in enumerate(order, start=1)}


def locate(lat, lon, code2pref):
    """座標の所在都道府県を地理院リバースジオコーダで引く。"""
    try:
        res = json.load(urllib.request.urlopen(
            GSI_REV.format(lat=lat, lon=lon), timeout=20))["results"]
    except Exception:
        return None
    return code2pref.get(str(res.get("muniCd")).zfill(5)[:2])


def hav(a, b, c, d):
    r = 6371000
    p = math.radians
    x = (math.sin(p(c - a) / 2) ** 2
         + math.cos(p(a)) * math.cos(p(c)) * math.sin(p(d - b) / 2) ** 2)
    return 2 * r * math.asin(math.sqrt(x))


def key_from_env():
    key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if not key:
        raise SystemExit("環境変数 GOOGLE_MAPS_API_KEY を設定してください")
    return key


def validate(args):
    """座標が既知の行で誤差を測る。新しいエリアに適用する前の確認用。"""
    key = key_from_env()
    rows, _, _ = read_rows(args.src, args.list_sheet)
    truth = [r for r in rows
             if str(r["acc"] or "") == args.acc and isinstance(r["lat"], (int, float))]
    if args.area:
        truth = [r for r in truth if r["area"] == args.area]
    if not truth:
        raise SystemExit(f"座標精度が '{args.acc}' の行がありません")

    def job(r):
        hit, why = search(r["name"], r["pref"], key)
        return r, hit, why

    kept, dropped = [], []
    with cf.ThreadPoolExecutor(max_workers=WORKERS) as ex:
        for r, hit, why in ex.map(job, truth):
            (kept if hit else dropped).append(
                (r, hav(r["lat"], r["lon"], hit[0], hit[1]) if hit else None, why))
    ds = sorted(d for _, d, _ in kept)
    print(f"照会 {len(truth)}件 → 名称一致で採用 {len(kept)} / 除外 {len(dropped)}")
    if ds:
        print(f"採用分の誤差: 中央値 {statistics.median(ds):.0f}m / 最大 {max(ds):.0f}m")
        for th in (100, 500, 1000, 5000):
            print(f"  <{th}m: {sum(1 for x in ds if x < th)}/{len(ds)}")
    return kept


def fill(args):
    key = key_from_env()
    rows, _, _ = read_rows(args.src, args.list_sheet)
    p2a = pref_to_area(rows)
    code2pref = muni_code_to_pref(rows)

    wb = load_workbook(args.list)
    report, filled, skipped = [], 0, 0

    for sheet in [s.strip() for s in args.sheets.split(",")]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(2, c).value or "").replace("\n", "")
            if h:
                hdr[h] = c
        for token in ("変電所名", "エリア", "緯度", "経度"):
            if token not in hdr:
                raise SystemExit(f"シート'{sheet}'に列'{token}'がありません")

        targets = []
        for r in range(3, ws.max_row + 1):
            name = ws.cell(r, hdr["変電所名"]).value
            if not name:
                continue
            if args.area and ws.cell(r, hdr["エリア"]).value != args.area:
                continue
            if ws.cell(r, hdr["緯度"]).value is not None and not args.overwrite:
                continue
            targets.append((r, name,
                            ws.cell(r, hdr["都道府県"]).value if "都道府県" in hdr else "",
                            ws.cell(r, hdr["エリア"]).value))

        def job(t):
            r, name, pref, area = t
            hit, why = search(name, pref, key)
            if not hit:
                return r, name, pref, area, None, why
            # 名称が一致していても別地方の同名変電所であることがあるので、
            # 座標の所在地を引き直して行の都道府県／エリアと突き合わせる
            got = locate(hit[0], hit[1], code2pref)
            if got is None:
                return r, name, pref, area, None, "所在地を確認できなかった"
            if pref:
                if got != pref:
                    return r, name, pref, area, None, f"所在地が{got}で都道府県と一致しない"
            elif p2a.get(got) != area:
                return (r, name, pref, area, None,
                        f"所在地が{got}（{p2a.get(got)}エリア）で行のエリアと一致しない")
            return r, name, pref, area, hit, "採用"

        with cf.ThreadPoolExecutor(max_workers=WORKERS) as ex:
            for r, name, pref, area, hit, why in ex.map(job, targets):
                no = ws.cell(r, hdr["No."]).value if "No." in hdr else None
                if not hit:
                    skipped += 1
                    report.append([sheet, r, no, name, pref, "", "", "不採用", why, ""])
                    continue
                lat, lon, gname, addr = hit
                ws.cell(r, hdr["緯度"]).value = round(lat, 6)
                ws.cell(r, hdr["経度"]).value = round(lon, 6)
                if "出典メモ" in hdr:
                    ws.cell(r, hdr["出典メモ"]).value = f"{SOURCE_LABEL} {gname}"
                for k in ("緯度", "経度", "出典メモ"):
                    if k in hdr:
                        ws.cell(r, hdr[k]).fill = FILLED_FILL
                filled += 1
                report.append([sheet, r, no, name, pref, round(lat, 6), round(lon, 6),
                               "採用", gname, addr])

    wb.save(args.out or args.list)
    print(f"記入 {filled}件 / 見送り {skipped}件")
    print(f"出力: {args.out or args.list}")
    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["シート", "行", "No.", "変電所名", "都道府県", "緯度", "経度",
                        "判定", "Google名称/理由", "住所"])
            w.writerows(report)
        print(f"レポート: {args.report}")
    if filled:
        print("※ 次に merge_coords.py で本体xlsxへ反映してください。")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    v = sub.add_parser("validate", help="座標が既知の行で誤差を測る")
    v.add_argument("--in", dest="src", required=True)
    v.add_argument("--list-sheet", default="全国変電所リスト")
    v.add_argument("--acc", default="公表資料(手動)", help="正解として使う座標精度")
    v.add_argument("--area")

    f = sub.add_parser("fill", help="優先リストの緯度・経度欄を埋める")
    f.add_argument("--in", dest="src", required=True, help="本体xlsx（都道府県→エリアの導出に使う）")
    f.add_argument("--list-sheet", default="全国変電所リスト")
    f.add_argument("--list", required=True)
    f.add_argument("--out")
    f.add_argument("--sheets", default="優先S_A,劣後S_A")
    f.add_argument("--area", help="このエリアだけに限定する")
    f.add_argument("--report")
    f.add_argument("--overwrite", action="store_true")

    args = ap.parse_args()
    validate(args) if args.cmd == "validate" else fill(args)


if __name__ == "__main__":
    main()
