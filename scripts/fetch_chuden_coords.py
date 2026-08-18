#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中部電力パワーグリッド「系統空容量マッピング」から変電所座標を取り出し、
優先リスト（list_missing_coords.py の出力）の中部エリア行に記入する。

https://gridmap.powergrid.chuden.co.jp/ は Leaflet 製で、変電所と送電線を
1本の GeoJSON（geo_data/KRSIH001）で配信している。座標は WGS84 の
[経度, 緯度] でそのまま入っているため投影変換は不要。

  {"type":"Feature",
   "properties":{"変電所番号":"1036","設備名称":"南武平町変電所（１）",
                 "電圧階級":"２７５ｋＶ", ...},
   "geometry":{"type":"Point","coordinates":[136.911265, 35.164986]}}

【取得が手作業になる理由】
gridmap.powergrid.chuden.co.jp は Azure Front Door が全パスに 403
（"The request is blocked."）を返す。ブラウザUAでも robots.txt でも同じで、
兄弟ホスト powergrid.chuden.co.jp は 200 なのでこのホスト固有の制限。
そのため取得はブラウザで行い、保存した HAR を extract に渡す。

使い方:
  # ① ブラウザのDevTools(Network)で記録し "Save all as HAR with content" で保存
  # ② HARからGeoJSONを取り出して保存（以降ネットワーク不要）
  python scripts/fetch_chuden_coords.py extract \
    --har gridmap.powergrid.chuden.co.jp.har --out chuden_cache/KRSIH001.geojson.gz

  # ③ 優先リストの中部エリア行に記入
  python scripts/fetch_chuden_coords.py fill \
    --geo chuden_cache/KRSIH001.geojson.gz \
    --list docs/座標未取得_優先リスト.xlsx --report chuden_fill_report.csv

  # ④ 本体xlsxへ反映（merge_coords.py 側の役割）
"""
import argparse
import csv
import gzip
import json
import math
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from fetch_substation_coords import normalize_name

GEO_PATH_RE = re.compile(r"/geo_data/KRSIH001$")
PARAM_PATH_RE = re.compile(r"/geo_data/KRSIH002$")
SOURCE_LABEL = "中部電力PG 系統空容量マッピング"
FILLED_FILL = PatternFill("solid", fgColor="DDEBF7")   # 青=調べた座標を記入済み
TARGET_AREA = "中部"

# 同一敷地の別バンク（「瑞穂変電所（１）」「（２）」等）は重心を採る。
# それを超えて離れているものは別施設の可能性があるので採用しない。
SAME_SITE_LIMIT_M = 500.0


def strip_trailer(text):
    """KRSIH001/002 は本体JSONの後ろに "&&20260817090439" が付く形式で返る。"""
    return re.sub(r"&&\d+\s*$", "", text).strip()


def load_geojson(path):
    op = gzip.open if str(path).endswith(".gz") else open
    with op(path, "rt", encoding="utf-8") as f:
        return json.loads(strip_trailer(f.read()))


def extract_from_har(har_path, out_path):
    """HARから geo_data/KRSIH001 のレスポンス本文を取り出して保存する。"""
    with open(har_path, encoding="utf-8") as f:
        har = json.load(f)
    hit, param = None, None
    for e in har["log"]["entries"]:
        url = e["request"]["url"].split("?")[0]
        if GEO_PATH_RE.search(url):
            hit = e
        elif PARAM_PATH_RE.search(url):
            param = e
    if hit is None:
        raise SystemExit(
            "HARに geo_data/KRSIH001 がありません。DevToolsを開いたままページを"
            "リロード（Ctrl+R）してから保存し直してください。"
        )
    content = hit["response"].get("content", {})
    text = content.get("text")
    if not text:
        raise SystemExit(
            "レスポンス本文がHARに含まれていません。"
            "'Save all as HAR with content' で保存してください。"
        )
    if content.get("encoding") == "base64":
        import base64
        text = base64.b64decode(text).decode("utf-8")
    data = json.loads(strip_trailer(text))

    # KRSIH002 の taishobi（例: "2026年08月13日時点"）を出典メモに載せるため取り込む
    if param is not None:
        ptext = param["response"].get("content", {}).get("text") or ""
        try:
            taishobi = json.loads(strip_trailer(ptext)).get("taishobi", "")
        except ValueError:
            taishobi = ""
        if taishobi:
            data["_taishobi"] = taishobi
            print(f"対象日: {taishobi}")

    feats = data.get("features", [])
    pts = [f for f in feats if f.get("geometry", {}).get("type") == "Point"]
    named = [f for f in pts if f.get("properties", {}).get("設備名称")]
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    op = gzip.open if out_path.endswith(".gz") else open
    with op(out_path, "wt", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"features {len(feats)}件（Point {len(pts)} / うち設備名称あり {len(named)}）")
    print(f"出力: {out_path}")


def dist_m(a, b):
    """2点間の概算距離(m)。同一敷地判定にしか使わないので簡易式で足りる。"""
    (la1, lo1), (la2, lo2) = a, b
    dx = (lo2 - lo1) * 111_320 * math.cos(math.radians((la1 + la2) / 2))
    dy = (la2 - la1) * 110_574
    return math.hypot(dx, dy)


def index_by_name(data):
    """照合キー → その名前を持つ Point の座標リスト。

    名称は「他社変電所」「発電所」には付かない（第三者情報を排除した公表物のため）。
    設備名称のあるものだけが当社設備で、これが照合対象になる。
    """
    idx = {}
    for f in data.get("features", []):
        if f.get("geometry", {}).get("type") != "Point":
            continue
        props = f.get("properties", {})
        name = props.get("設備名称")
        if not name:
            continue
        key = normalize_name(name)
        if not key:
            continue
        lon, lat = f["geometry"]["coordinates"][:2]
        idx.setdefault(key, []).append({
            "lat": lat, "lon": lon, "name": name,
            "voltage": props.get("電圧階級", ""),
            "no": props.get("変電所番号", ""),
        })
    return idx


def resolve(cands):
    """同名候補から採用座標を決める。(lat, lon, spread_m, 理由) を返す。"""
    pts = [(c["lat"], c["lon"]) for c in cands]
    spread = max((dist_m(a, b) for a in pts for b in pts), default=0.0)
    if spread > SAME_SITE_LIMIT_M:
        return None, None, spread, f"同名{len(cands)}件が{spread:.0f}m離れており同一敷地と見なせない"
    lat = sum(p[0] for p in pts) / len(pts)
    lon = sum(p[1] for p in pts) / len(pts)
    return lat, lon, spread, "採用"


def fill(args):
    data = load_geojson(args.geo)
    idx = index_by_name(data)
    taishobi = data.get("_taishobi", "")
    wb = load_workbook(args.list)
    report, filled, skipped = [], 0, 0

    for sheet in [s.strip() for s in args.sheets.split(",")]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(2, c).value or "").replace("\n", "")
            if h:
                hdr[h] = c
        for token in ("エリア", "変電所名", "緯度", "経度"):
            if token not in hdr:
                raise SystemExit(f"シート'{sheet}'に列'{token}'がありません: {list(hdr)}")

        for r in range(3, ws.max_row + 1):
            area = ws.cell(r, hdr["エリア"]).value
            # 中部エリアの行だけを対象にする。gridmapの名前は他エリアの同名変電所とも
            # 一致してしまう（このリスト内で60行）ため、エリアを限定しないと誤適用になる。
            if area != TARGET_AREA:
                continue
            name = ws.cell(r, hdr["変電所名"]).value
            no = ws.cell(r, hdr["No."]).value if "No." in hdr else None
            if ws.cell(r, hdr["緯度"]).value is not None and not args.overwrite:
                skipped += 1
                report.append([sheet, r, no, name, "", "", "不採用", "既に記入がある"])
                continue
            cands = idx.get(normalize_name(name))
            if not cands:
                skipped += 1
                report.append([sheet, r, no, name, "", "", "不採用",
                               "gridmapに同名の当社設備が無い"])
                continue
            lat, lon, spread, why = resolve(cands)
            if lat is None:
                skipped += 1
                report.append([sheet, r, no, name, "", "", "不採用", why])
                continue
            ws.cell(r, hdr["緯度"]).value = round(lat, 6)
            ws.cell(r, hdr["経度"]).value = round(lon, 6)
            memo = SOURCE_LABEL + (f" {taishobi}" if taishobi else "")
            if len(cands) > 1:
                memo += f"（同名{len(cands)}件の重心・最大{spread:.0f}m）"
            if "出典メモ" in hdr:
                ws.cell(r, hdr["出典メモ"]).value = memo
            for key in ("緯度", "経度", "出典メモ"):
                if key in hdr:
                    ws.cell(r, hdr[key]).fill = FILLED_FILL
            filled += 1
            report.append([sheet, r, no, name, round(lat, 6), round(lon, 6), "採用", memo])

    wb.save(args.out or args.list)
    print(f"中部エリア: 記入 {filled}件 / 見送り {skipped}件")
    print(f"出力: {args.out or args.list}")
    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["シート", "行", "No.", "変電所名", "緯度", "経度", "判定", "理由"])
            w.writerows(report)
        print(f"レポート: {args.report}")
    if filled:
        print("※ 次に merge_coords.py で本体xlsxへ反映してください。")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    e = sub.add_parser("extract", help="HARからGeoJSONを取り出して保存する")
    e.add_argument("--har", required=True)
    e.add_argument("--out", default="chuden_cache/KRSIH001.geojson.gz")

    f = sub.add_parser("fill", help="優先リストの中部エリア行に座標を記入する")
    f.add_argument("--geo", default="chuden_cache/KRSIH001.geojson.gz")
    f.add_argument("--list", required=True, help="優先リストのxlsx")
    f.add_argument("--out", help="書き出し先（既定は --list を上書き）")
    f.add_argument("--sheets", default="優先S_A,劣後S_A")
    f.add_argument("--report")
    f.add_argument("--overwrite", action="store_true",
                   help="既に記入がある行も上書きする")

    args = ap.parse_args()
    if args.cmd == "extract":
        extract_from_har(args.har, args.out)
    else:
        fill(args)


if __name__ == "__main__":
    main()
